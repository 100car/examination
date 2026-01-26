"""
STEP 8 | Forecast ring availability (Excel formulas)

Модуль додає на лист RINGS_SUMMARY прогноз дефіциту поршневих кілець (SKU)
на N місяців вперед у вигляді формул Excel.

Чому саме формули:
- якщо в *-normalized зміняться Month_sales / Stock, прогноз автоматично перераховується у Excel;
- зручно для демонстрації в дипломі (видно прозору бізнес-логіку без “чорної скриньки”).

Логіка (для кожного SKU, кожного місяця m, і кожного регіону UA/KZ/UZ):
  deficit_region = m * Month_sales - Stock
  need_region_m  = IF(deficit_region > 0, -deficit_region, 0)

Потім підсумовуємо 3 регіони і враховуємо Total Available з RINGS_SUMMARY:
  need_total_m = MIN(0, TotalAvailable + need_UA_m + need_KZ_m + need_UZ_m)

Інтерпретація:
- 0  → дефіциту немає
- <0 → дефіцит (скільки не вистачає)

Додатково:
- заголовки NEED_M* робляться вертикальними (аналогічно Order_1)
- для значень < 0 застосовується умовне форматування (світло‑червона заливка + темно‑червоний текст)
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule


# =============================================================================
# ШЛЯХИ ПРОЄКТУ (привʼязані до місця файлу у папці DATA/)
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent              # .../DATA
SALES_ROOT_DIR = BASE_DIR / "SALES"                    # .../DATA/SALES
RESULT_DIR = SALES_ROOT_DIR / "RESULT"                 # .../DATA/SALES/RESULT
DEFAULT_RESULTS_XLSX = RESULT_DIR / "results.xlsx"     # .../DATA/SALES/RESULT/results.xlsx


# Excel number format (₽) як у вихідному прикладі
_NUMBER_FORMAT_RUB = r'_-* # ##0\ _₽_-;-* # ##0\ _₽_-;_-* "-"??\ _₽_-;_-@_-'


def _print_banner(text: str) -> None:
    line = "=" * max(12, len(text))
    print(f"\n{line}\n{text}\n{line}")


@dataclass(frozen=True)
class RegionSpec:
    """Опис того, де шукати Month_sales / Stock / SKU на листі регіону."""

    sheet_name: str
    month_sales_col: str = "D"  # Month_sales
    stock_col: str = "E"       # Stock
    sku_col: str = "A"         # Symbol_normalized (або Symbol)


def _find_col_by_header(ws, header_value: str) -> Optional[int]:
    """Повертає номер колонки (1-based) за значенням заголовка у 1-му рядку."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == header_value:
            return c
    return None


def _delete_columns_from(ws, start_col: int) -> None:
    """Видаляє всі колонки починаючи зі start_col до кінця (включно)."""
    if start_col <= ws.max_column:
        ws.delete_cols(start_col, ws.max_column - start_col + 1)


def _region_need_formula(region: RegionSpec, sku_cell: str, month: int) -> str:
    """Формула дефіциту по одному регіону для SKU у sku_cell на заданий month (1..N)."""
    ms = (
        f"INDEX('{region.sheet_name}'!${region.month_sales_col}:${region.month_sales_col}, "
        f"MATCH({sku_cell}, '{region.sheet_name}'!${region.sku_col}:${region.sku_col}, 0))"
    )
    st = (
        f"INDEX('{region.sheet_name}'!${region.stock_col}:${region.stock_col}, "
        f"MATCH({sku_cell}, '{region.sheet_name}'!${region.sku_col}:${region.sku_col}, 0))"
    )
    return f"IFERROR(IF({month}*{ms}-{st}>0, -({month}*{ms}-{st}), 0), 0)"


def generate_rings_forecast(results_path: str | Path, forecast_months: int = 30) -> None:
    """Додає колонки NEED_M1..NEED_MN на RINGS_SUMMARY як формули Excel + форматування."""
    results_path = Path(results_path)

    _print_banner(f"STEP 8 | Прогноз наявності поршневих кілець | months={forecast_months}")
    if not results_path.exists():
        raise FileNotFoundError(f"results.xlsx не знайдено: {results_path}")

    wb = load_workbook(results_path)
    if "RINGS_SUMMARY" not in wb.sheetnames:
        raise ValueError('Не знайдено лист "RINGS_SUMMARY" у results.xlsx. Спочатку запусти генерацію RINGS_SUMMARY.')

    ws = wb["RINGS_SUMMARY"]

    total_col = _find_col_by_header(ws, "Total Available")
    if not total_col:
        raise ValueError('Не знайдено заголовок "Total Available" на листі RINGS_SUMMARY.')

    _delete_columns_from(ws, total_col + 1)
    start_col = total_col + 1

    order_col = _find_col_by_header(ws, "Order_1")
    header_alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
    header_font = Font(bold=True)

    if order_col:
        oc = ws.cell(row=1, column=order_col)
        if oc.alignment:
            header_alignment = Alignment(
                horizontal=oc.alignment.horizontal or "center",
                vertical=oc.alignment.vertical or "center",
                textRotation=oc.alignment.textRotation or 90,
                wrap_text=oc.alignment.wrap_text,
            )
        if oc.font:
            header_font = Font(
                name=oc.font.name,
                size=oc.font.size,
                bold=True,
                italic=oc.font.italic,
                underline=oc.font.underline,
                color=oc.font.color,
            )

    for m in range(1, forecast_months + 1):
        cell = ws.cell(row=1, column=start_col + (m - 1))
        cell.value = f"NEED_M{m}"
        cell.alignment = header_alignment
        cell.font = header_font

    regions = [RegionSpec("UA-normalized"), RegionSpec("KZ-normalized"), RegionSpec("UZ-normalized")]
    max_row = ws.max_row

    for r in range(2, max_row + 1):
        sku_val = ws.cell(row=r, column=1).value
        if sku_val is None or str(sku_val).strip() == "":
            continue

        sku_cell = f"$A{r}"
        total_cell = ws.cell(row=r, column=total_col).coordinate

        for m in range(1, forecast_months + 1):
            ua = _region_need_formula(regions[0], sku_cell, m)
            kz = _region_need_formula(regions[1], sku_cell, m)
            uz = _region_need_formula(regions[2], sku_cell, m)
            ws.cell(row=r, column=start_col + (m - 1)).value = f"=MIN(0, {total_cell}+({ua})+({kz})+({uz}))"

    first_data_row = 2
    last_data_row = max_row
    first_col_letter = ws.cell(row=1, column=start_col).column_letter
    last_col_letter = ws.cell(row=1, column=start_col + forecast_months - 1).column_letter
    data_range = f"{first_col_letter}{first_data_row}:{last_col_letter}{last_data_row}"

    for c in range(start_col, start_col + forecast_months):
        for rr in range(first_data_row, last_data_row + 1):
            ws.cell(row=rr, column=c).number_format = _NUMBER_FORMAT_RUB

    fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    font = Font(color="9C0006")
    rule = CellIsRule(operator="lessThan", formula=["0"], fill=fill, font=font)
    ws.conditional_formatting.add(data_range, rule)

    wb.save(results_path)
    print(f"OK | updated: {results_path} | forecast_months={forecast_months}")


def _build_cli() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="STEP 8: Додати NEED_M* прогноз (Excel формули) на RINGS_SUMMARY.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--results", type=str, default=str(DEFAULT_RESULTS_XLSX), help="Шлях до results.xlsx")
    p.add_argument("--months", type=int, default=30, help="Горизонт прогнозу (місяці)")
    return p


def main() -> None:
    args = _build_cli().parse_args()
    generate_rings_forecast(args.results, forecast_months=args.months)


if __name__ == "__main__":
    main()
