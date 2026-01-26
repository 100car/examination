from __future__ import annotations
"""
!!!! Перевірити як комплекти рахує!!!!

rings_summary.py

Утиліта формування та заповнення листа RINGS_SUMMARY у results.xlsx.

STEP 5a. Побудувати/перебудувати лист RINGS_SUMMARY (перший лист) з sku_master.xlsx:
- беремо рядки, де Name == 1
- пишемо колонки: Symbol, Brand, Group1, Group2

STEP 5b. Заповнити RINGS_SUMMARY даними з workbook *sets* (SALES/TRANSIT_STOCK_STATUS):
- для кожного листа sets -> окрема колонка у RINGS_SUMMARY (назва колонки = назва листа)
- матчинг по SKU + нормалізація + Mapping (Неправильный -> Артикул)

STEP 5c. Перевірка сум для кожної sets-колонки:
- sum_sets (у листі sets) vs sum_results (у відповідній колонці RINGS_SUMMARY)
- допуск TOL = 0.0001

STEP 5d. Заповнити RINGS_SUMMARY з surowce.xlsx:
- у кожному листі surowce є 2 колонки:
    A: "Номер по каталогу"
    B: "В С Е Г О"
- рядки бувають 2-х видів:
    * "неспрощені": A починається на "PR-" => Group2 = ПСТР(A;8;11) (у Python: A[7:18])
    * "спрощені":   інакше Group2 вже записаний у A (типу "14-2802-000")
- для кожного Group2 беремо MIN(B)
- далі через sku_master (Name==1) мапимо Group2 -> (SKU, AssemblyQty)
    * якщо однаковий Group2 у кількох SKU — беремо той, у кого найменший AssemblyQty
- записуємо у RINGS_SUMMARY: Q-ty = MIN(B) / AssemblyQty

Особливий кейс Group2 == "19-1402-000":
- замість sku_master створюємо 2 SKU:
    PR-DAE-39-1402-000-SET та PR-DAE-49-1402-000-SET
- ваги беремо з results.xlsx, лист "UZ-normalized", колонка TOTAL
- розподіл:
    q39 = (q * w39/(w39*AssemblyQty39+w49*AssemblyQty49)) / AssemblyQty39  
    q49 = (q * w49/(w39*AssemblyQty39+w49*AssemblyQty49)) / AssemblyQty49
    (AssemblyQty39 = 3, AssemblyQty49 = 4 - беремо з файлу sku_master.xlsx)
  де q — це MIN(B) для Group2 == "19-1402-000"

STEP 5e. Додаткова колонка, якщо у results.xlsx є лист "STOCKS-":
- створюємо наступну колонку (після "STOCKS-") з назвою "STOCKS-"
- пишемо значення з ПРОТИЛЕЖНИМ ЗНАКОМ (тобто -Quantity)
- перевіряємо суму: sum_neg має дорівнювати -sum_sets (TOL)

Формат консолі:
- для цілих чисел — без дробової частини
- для нецілих — 1 знак після коми
"""

import datetime
import os
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
SALES_ROOT_DIR = BASE_DIR / "SALES"

# results.xlsx
RESULT_DIR = SALES_ROOT_DIR / "RESULT"
DEFAULT_RESULTS_XLSX = str(RESULT_DIR / "results.xlsx")

# sku_master.xlsx
SKU_ROOT_DIR = BASE_DIR / "SKU"
DEFAULT_SKU_MASTER_XLSX = str(SKU_ROOT_DIR / "sku_master.xlsx")

# transit папки
DEFAULT_TRANSIT_STOCK_STATUS_DIR = str(SALES_ROOT_DIR / "TRANSIT_STOCK_STATUS")
DEFAULT_TRANSIT_PLANNED_ARRIVALS_DIR = str(SALES_ROOT_DIR / "TRANSIT_PLANNED_ARRIVALS")

SHEET_SKU = "SKU"
SHEET_MAPPING = "Mapping"
SHEET_RINGS_SUMMARY = "RINGS_SUMMARY"
SHEET_UZ_NORM = "UZ-normalized"

HDR_SKU = "SKU (ключ, унікальний)"
HDR_NAME_FLAG = "Name"
HDR_BRAND = "Brand"
HDR_GROUP1 = "Group1"
HDR_GROUP2 = "Group2"
HDR_ASSEMBLY_QTY = "AssemblyQty"
HDR_MAP_BAD = "Неправильный"
HDR_MAP_GOOD = "Артикул"

TOL = 0.0001


# --------------------------- форматування / консоль ---------------------------

def _fmt_status(tag: str, use_color: bool = True) -> str:
    if not use_color:
        return tag
    colors = {"OK": "\x1b[32m", "WRN": "\x1b[33m", "ERR": "\x1b[31m"}
    reset = "\x1b[0m"
    return f"{colors.get(tag, '')}{tag}{reset}"


def _print_banner(title: str) -> None:
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = "=" * 72
    print(line)
    print(f"{title} | {ts}")
    print(line)


def _is_int_like(x: float, tol: float = 1e-9) -> bool:
    return abs(x - round(x)) <= tol


def _fmt_num(x: float) -> str:
    """Цілі -> без десяткових; дробові -> 1 знак після коми."""
    if _is_int_like(x):
        return f"{int(round(x))}"
    return f"{x:.1f}"


# --------------------------- базові утиліти ---------------------------

def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except Exception:
        return None


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    try:
        return float(s.replace(" ", "").replace("\u00A0", "").replace(",", "."))
    except Exception:
        return 0.0


def _find_header_col(ws, header: str) -> Optional[int]:
    header = header.strip()
    for c in range(1, ws.max_column + 1):
        if _norm_cell(ws.cell(1, c).value) == header:
            return c
    return None


def _norm_header(v: Any) -> str:
    s = _norm_cell(v).lower().replace("\n", " ").replace("\r", " ")
    return " ".join(s.split())


def _find_header_col_any(ws, candidates: List[str]) -> Optional[int]:
    cand = {_norm_header(c) for c in candidates}
    for c in range(1, ws.max_column + 1):
        if _norm_header(ws.cell(1, c).value) in cand:
            return c
    return None


_C2L = {
    "А": "A", "В": "B", "С": "C", "Е": "E", "Н": "H", "К": "K", "М": "M", "О": "O",
    "Р": "P", "Т": "T", "Х": "X", "а": "a", "в": "b", "с": "c", "е": "e", "н": "h",
    "к": "k", "м": "m", "о": "o", "р": "p", "т": "t", "х": "x",
}


def _normalize_sku(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", "").replace("\u202F", "")
    s = re.sub(r"\s+", "", s)
    s = "".join(_C2L.get(ch, ch) for ch in s)
    return s.strip()


def _autosize(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 0
        for row in range(1, min(ws.max_row, 5000) + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[letter].width = min(max(best + 2, 10), max_width)


def _rotate_headers_vertical(ws, start_col: int = 5) -> None:
    """
    Повертаємо заголовки (рядок 1) вертикально для "даних" колонок,
    щоб назви листів поміщались.
    За замовчуванням: з колонки 5 (після Symbol/Brand/Group1/Group2).
    """
    al = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)
    for c in range(start_col, ws.max_column + 1):
        cell = ws.cell(1, c)
        if cell.value is None:
            continue
        cell.alignment = al
    ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 15, 80)


def _center_columns(ws, cols: List[int], start_row: int = 1, end_row: Optional[int] = None) -> None:
    """Вирівнюємо по центру (горизонтально + вертикально) в заданих колонках."""
    al = Alignment(horizontal="center", vertical="center", wrap_text=False)
    if end_row is None:
        end_row = ws.max_row
    for r in range(start_row, end_row + 1):
        for c in cols:
            ws.cell(r, c).alignment = al


def _move_column_after(ws, col_name: str, after_col_name: str) -> None:
    """
    Переміщує колонку з заголовком col_name одразу після after_col_name (рядок 1).
    Якщо чогось немає — нічого не робить.
    """
    header_to_col: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = _norm_cell(ws.cell(1, c).value)
        if h:
            header_to_col[h] = c
    if col_name not in header_to_col or after_col_name not in header_to_col:
        return
    col_from = header_to_col[col_name]
    col_after = header_to_col[after_col_name]
    if col_from == col_after + 1:
        return

    # Беремо весь стовбчик як список значень + стилів
    max_row = ws.max_row
    values = [ws.cell(r, col_from).value for r in range(1, max_row + 1)]
    styles = [ws.cell(r, col_from)._style for r in range(1, max_row + 1)]
    number_formats = [ws.cell(r, col_from).number_format for r in range(1, max_row + 1)]
    alignments = [ws.cell(r, col_from).alignment for r in range(1, max_row + 1)]
    fonts = [ws.cell(r, col_from).font for r in range(1, max_row + 1)]

    # Видаляємо колонку
    ws.delete_cols(col_from, 1)

    # Переобчислюємо позицію after (якщо видаляли колонку лівіше)
    if col_from < col_after:
        col_after -= 1
    insert_at = col_after + 1

    ws.insert_cols(insert_at, 1)

    # Відновлюємо
    for r in range(1, max_row + 1):
        cell = ws.cell(r, insert_at)
        cell.value = values[r - 1]
        cell._style = styles[r - 1]
        cell.number_format = number_formats[r - 1]
        cell.alignment = alignments[r - 1]
        cell.font = fonts[r - 1]

def add_orders_and_total_available(ws_sum) -> None:
    """
    STEP 5g (фінальний):

    1) Додаємо 3 порожні колонки: Order_1, Order_2, Order_3.
    2) Праворуч додаємо колонку 'Total Available'.
    3) У 'Total Available' записуємо ФОРМУЛУ сумування всіх колонок зліва (для цього рядка).

    Примітка:
    - SUM() ігнорує текстові значення, але ми свідомо рахуємо лише "дані" колонки,
      починаючи з 5-ї (після Group2).
    """
    _ensure_column(ws_sum, "Order_1")
    _ensure_column(ws_sum, "Order_2")
    _ensure_column(ws_sum, "Order_3")

    total_col = _ensure_column(ws_sum, "Total Available")

    first_data_col = 5  # після Symbol/Brand/Group1/Group2
    last_left_col = total_col - 1
    if last_left_col < first_data_col:
        return

    c1 = get_column_letter(first_data_col)
    c2 = get_column_letter(last_left_col)

    for r in range(2, ws_sum.max_row + 1):
        ws_sum.cell(r, total_col, f"=SUM({c1}{r}:{c2}{r})")



# --------------------------- sku_master / mapping ---------------------------

def _load_mapping(sku_master_path: str) -> Dict[str, str]:
    """normalized 'Неправильный' -> normalized 'Артикул'"""
    try:
        wb = openpyxl.load_workbook(sku_master_path, data_only=True)
    except Exception:
        return {}
    if SHEET_MAPPING not in wb.sheetnames:
        return {}
    ws = wb[SHEET_MAPPING]
    col_bad = _find_header_col(ws, HDR_MAP_BAD)
    col_good = _find_header_col(ws, HDR_MAP_GOOD)
    if col_bad is None or col_good is None:
        return {}
    mp: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        bad = _normalize_sku(ws.cell(r, col_bad).value)
        good = _normalize_sku(ws.cell(r, col_good).value)
        if bad and good:
            mp[bad] = good
    return mp


def _load_group2_to_best_sku(sku_master_path: str) -> Dict[str, Tuple[str, float]]:
    """
    Для рядків Name==1:
      Group2 -> (SKU, AssemblyQty), де при дублях Group2 беремо мінімальний AssemblyQty.
    """
    sku_wb = openpyxl.load_workbook(sku_master_path, data_only=True)
    if SHEET_SKU not in sku_wb.sheetnames:
        raise ValueError(f"sku_master.xlsx: лист '{SHEET_SKU}' не знайдено")
    ws = sku_wb[SHEET_SKU]

    col_sku = _find_header_col(ws, HDR_SKU)
    col_name = _find_header_col(ws, HDR_NAME_FLAG)
    col_g2 = _find_header_col(ws, HDR_GROUP2)
    col_asm = _find_header_col_any(ws, [HDR_ASSEMBLY_QTY, "Assembly Qty", "AsmQty", "Qty per set"])

    missing = [h for h, c in [(HDR_SKU, col_sku), (HDR_NAME_FLAG, col_name), (HDR_GROUP2, col_g2), (HDR_ASSEMBLY_QTY, col_asm)] if c is None]
    if missing:
        raise ValueError(f"sku_master.xlsx: бракує колонок: {missing}")

    out: Dict[str, Tuple[str, float]] = {}
    for r in range(2, ws.max_row + 1):
        if _to_int(ws.cell(r, col_name).value) != 1:
            continue
        sku = _norm_cell(ws.cell(r, col_sku).value)
        g2 = _norm_cell(ws.cell(r, col_g2).value)
        asm = _to_float(ws.cell(r, col_asm).value)
        if not sku or not g2:
            continue
        if asm <= 0:
            asm = 1.0
        if g2 not in out:
            out[g2] = (sku, asm)
        else:
            _, prev_asm = out[g2]
            if asm < prev_asm:
                out[g2] = (sku, asm)
    return out


# --------------------------- побудова RINGS_SUMMARY ---------------------------

def build_rings_summary_sheet(
    results_xlsx_path: str = DEFAULT_RESULTS_XLSX,
    sku_master_path: str = DEFAULT_SKU_MASTER_XLSX,
    *,
    use_color: bool = True,
) -> str:
    """Створити/перебудувати лист RINGS_SUMMARY (перший лист) з sku_master (Name==1)."""
    _print_banner("STEP 5a | Побудова RINGS_SUMMARY")

    sku_wb = openpyxl.load_workbook(sku_master_path, data_only=True)
    if SHEET_SKU not in sku_wb.sheetnames:
        raise ValueError(f"sku_master.xlsx: лист '{SHEET_SKU}' не знайдено")
    sku_ws = sku_wb[SHEET_SKU]

    col_sku = _find_header_col(sku_ws, HDR_SKU)
    col_name = _find_header_col(sku_ws, HDR_NAME_FLAG)
    col_brand = _find_header_col(sku_ws, HDR_BRAND)
    col_g1 = _find_header_col(sku_ws, HDR_GROUP1)
    col_g2 = _find_header_col(sku_ws, HDR_GROUP2)

    missing_cols = [h for h, c in [(HDR_SKU, col_sku), (HDR_NAME_FLAG, col_name), (HDR_BRAND, col_brand), (HDR_GROUP1, col_g1), (HDR_GROUP2, col_g2)] if c is None]
    if missing_cols:
        raise ValueError(f"sku_master.xlsx: бракує колонок: {missing_cols}")

    rows: List[Tuple[str, str, str, str]] = []
    for r in range(2, sku_ws.max_row + 1):
        if _to_int(sku_ws.cell(r, col_name).value) != 1:
            continue
        sku = _norm_cell(sku_ws.cell(r, col_sku).value)
        if not sku:
            continue
        brand = _norm_cell(sku_ws.cell(r, col_brand).value)
        g1 = _norm_cell(sku_ws.cell(r, col_g1).value)
        g2 = _norm_cell(sku_ws.cell(r, col_g2).value)
        rows.append((sku, brand, g1, g2))

    rows.sort(key=lambda t: (t[1], t[2], t[3], t[0]))

    wb = openpyxl.load_workbook(results_xlsx_path)
    if SHEET_RINGS_SUMMARY in wb.sheetnames:
        wb.remove(wb[SHEET_RINGS_SUMMARY])
    ws = wb.create_sheet(SHEET_RINGS_SUMMARY, 0)

    headers = ["Symbol", "Brand", "Group1", "Group2"]
    bold = Font(bold=True)
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h).font = bold

    for i, (sku, brand, g1, g2) in enumerate(rows, start=2):
        ws.cell(i, 1, sku)
        ws.cell(i, 2, brand)
        ws.cell(i, 3, g1)
        ws.cell(i, 4, g2)

    ws.freeze_panes = "A2"
    _center_columns(ws, cols=[1, 2, 3, 4], start_row=1)
    _autosize(ws)

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)} | step=5a | sheet={SHEET_RINGS_SUMMARY} | rows={len(rows)} | saved={os.path.abspath(results_xlsx_path)}")
    return results_xlsx_path


# --------------------------- sets: читання, заповнення, перевірка ---------------------------

def _find_sets_file(transit_dir: str) -> Optional[str]:
    if not transit_dir or not os.path.isdir(transit_dir):
        return None
    files = [f for f in os.listdir(transit_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    sets = [f for f in files if "sets" in f.lower()]
    pick = sorted(sets or files)
    return os.path.join(transit_dir, pick[0]) if pick else None


def _read_sheet_sku_qty(ws) -> Dict[str, float]:
    sku_col = _find_header_col_any(ws, ["SKU", "Symbol", "Артикул", "SKU (ключ, унікальний)"])
    qty_col = _find_header_col_any(ws, ["Quantity", "Qty", "Кількість", "Количество", "Остаток", "Остатки", "Stock", "заказ"])
    if sku_col is None or qty_col is None:
        raise ValueError(f"Лист '{ws.title}': не знайдено колонки SKU/Quantity")
    out: Dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        sku = _norm_cell(ws.cell(r, sku_col).value)
        if not sku:
            continue
        out[sku] = out.get(sku, 0.0) + _to_float(ws.cell(r, qty_col).value)
    return out


def _ensure_column(ws_sum, header: str) -> int:
    for c in range(1, ws_sum.max_column + 1):
        if _norm_cell(ws_sum.cell(1, c).value) == header:
            return c
    c = ws_sum.max_column + 1
    ws_sum.cell(1, c, header).font = Font(bold=True)
    return c


def _build_summary_row_index(ws_sum) -> Tuple[Dict[str, int], Dict[str, int]]:
    sym_to_row: Dict[str, int] = {}
    sym_norm_to_row: Dict[str, int] = {}
    for r in range(2, ws_sum.max_row + 1):
        sym = _norm_cell(ws_sum.cell(r, 1).value)
        if sym:
            sym_to_row[sym] = r
            sym_norm_to_row[_normalize_sku(sym)] = r
    return sym_to_row, sym_norm_to_row


def apply_transit_stock_sets(
    results_xlsx_path: str = DEFAULT_RESULTS_XLSX,
    *,
    transit_dir: str = DEFAULT_TRANSIT_STOCK_STATUS_DIR,
    sets_path: Optional[str] = None,
    surowce_path: Optional[str] = None,
    sku_master_path: str = DEFAULT_SKU_MASTER_XLSX,
    apply_surowce: bool = True,
    use_color: bool = True,
) -> str:
    """
    Головна процедура: sets (5b/5c) + (опційно) surowce (5d).
    """
    _print_banner("STEP 5b | Заповнення з sets + перевірка сум (5c)")

    sets_file = sets_path or _find_sets_file(transit_dir)
    if not sets_file:
        print(f"{_fmt_status('WRN', use_color)} | step=5b | reason=missing_sets_file | dir={os.path.abspath(transit_dir)}")
        if apply_surowce:
            return apply_surowce_workbook(
                results_xlsx_path,
                transit_dir=transit_dir,
                surowce_path=surowce_path,
                sku_master_path=sku_master_path,
                use_color=use_color,
            )
        return results_xlsx_path

    wb = openpyxl.load_workbook(results_xlsx_path)
    if SHEET_RINGS_SUMMARY not in wb.sheetnames:
        raise ValueError(f"results workbook: немає '{SHEET_RINGS_SUMMARY}'. Спочатку запусти build_rings_summary_sheet().")
    ws_sum = wb[SHEET_RINGS_SUMMARY]

    sym_to_row, sym_norm_to_row = _build_summary_row_index(ws_sum)
    if not sym_to_row:
        raise ValueError("RINGS_SUMMARY порожній (немає рядків).")

    mapping = _load_mapping(sku_master_path)
    sets_wb = openpyxl.load_workbook(sets_file, data_only=True)

    mapped_total = 0
    missing_total = 0

    for sheet_name in sets_wb.sheetnames:
        ws_sets = sets_wb[sheet_name]
        try:
            mp = _read_sheet_sku_qty(ws_sets)
        except Exception as exc:
            print(f"{_fmt_status('WRN', use_color)} | step=5b | sheet={sheet_name} | skipped | {exc}")
            continue

        col_main = _ensure_column(ws_sum, sheet_name)

        filled = 0
        mapped = 0
        missing = 0

        sum_sets = float(sum(mp.values()))
        sum_results = 0.0

        for sku_raw, qty in mp.items():
            rr = sym_to_row.get(sku_raw)
            if rr is None:
                rr = sym_norm_to_row.get(_normalize_sku(sku_raw))
            if rr is None:
                mapped_sku = mapping.get(_normalize_sku(sku_raw))
                if mapped_sku:
                    rr = sym_norm_to_row.get(mapped_sku)
                    if rr is not None:
                        mapped += 1
                        mapped_total += 1

            if rr is None:
                missing += 1
                missing_total += 1
                continue

            ws_sum.cell(rr, col_main, qty)
            sum_results += qty

            filled += 1

        diff = abs(sum_sets - sum_results)
        status = "OK" if diff <= TOL else "WRN"

        msg = (
            f"{_fmt_status(status, use_color)} | step=5c | sheet={sheet_name:<20} | in={len(mp):<6} | "
            f"filled={filled:<6} | mapped={mapped:<6} | missing={missing:<6} | "
            f"sum_sets={_fmt_num(sum_sets):>8} | sum_results={_fmt_num(sum_results):>8} | diff={_fmt_num(diff):>8}"
        )

        print(msg)

    # Зробимо RINGS_SUMMARY першим листом
    if wb.sheetnames[0] != SHEET_RINGS_SUMMARY:
        wb._sheets.sort(key=lambda s: 0 if s.title == SHEET_RINGS_SUMMARY else 1)

    # Вертикальні заголовки для даних колонок (починаючи з 5-ї)
    _rotate_headers_vertical(ws_sum, start_col=5)
    # Центруємо базові колонки
    _center_columns(ws_sum, cols=[1, 2, 3, 4], start_row=1)

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)} | step=5b | saved={os.path.abspath(results_xlsx_path)} | mapped_total={mapped_total} | missing_total={missing_total}")

    if apply_surowce:
        apply_surowce_workbook(
            results_xlsx_path,
            transit_dir=transit_dir,
            surowce_path=surowce_path,
            sku_master_path=sku_master_path,
            use_color=use_color,
        )

    return results_xlsx_path


# --------------------------- surowce: читання, заповнення, перевірка ---------------------------

def _find_first_xlsx(transit_dir: str, name_contains: str) -> Optional[str]:
    if not transit_dir or not os.path.isdir(transit_dir):
        return None
    files = [f for f in os.listdir(transit_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    picks = [f for f in files if name_contains.lower() in f.lower()]
    pick = sorted(picks or [])
    return os.path.join(transit_dir, pick[0]) if pick else None

def _list_xlsx_files(dir_path: str) -> List[str]:
    """Повертає всі *.xlsx у папці (без тимчасових файлів)."""
    if not dir_path or not os.path.isdir(dir_path):
        return []
    files = [
        os.path.join(dir_path, f)
        for f in os.listdir(dir_path)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]
    return sorted(files)


def _read_headerless_group2_qty_min(ws) -> Dict[str, float]:
    """
    Для листів warehouse_arrivals/TRANSIT_PLANNED_ARRIVALS без рядка заголовків:
    - Колонка A: Group2 (типу '19-1402-000')
    - Колонка B: Quantity
    Повертає {Group2: MIN(Quantity)} (як і в surowce).
    """
    out: Dict[str, float] = {}
    rows_seen = 0
    for r in range(1, ws.max_row + 1):
        g2 = _norm_cell(ws.cell(r, 1).value)
        q = _to_float(ws.cell(r, 2).value)
        if not g2 and q == 0.0:
            continue
        if not g2:
            continue
        rows_seen += 1
        if g2 not in out:
            out[g2] = q
        else:
            out[g2] = min(out[g2], q)
    if rows_seen == 0:
        raise ValueError(f"Лист '{ws.title}': немає даних у колонках A/B")
    return out


def apply_planned_arrivals_workbooks(
    results_xlsx_path: str,
    *,
    planned_dir: str = DEFAULT_TRANSIT_PLANNED_ARRIVALS_DIR,
    planned_paths: Optional[List[str]] = None,
    sku_master_path: str = DEFAULT_SKU_MASTER_XLSX,
    use_color: bool = True,
) -> str:
    """
    STEP 5f: TRANSIT_PLANNED_ARRIVALS (warehouse_arrivals тощо).

    - Беремо ВСІ *.xlsx з папки SALES/TRANSIT_PLANNED_ARRIVALS (може бути 1 або більше)
    - У кожному файлі: кожен лист -> окрема колонка в RINGS_SUMMARY
    - Формат листів: БЕЗ заголовків (A=Group2, B=Quantity)
    - Логіка така сама як для surowce (включно з розподілом 19-1402-000).
    """
    _print_banner("STEP 5f | Заповнення з TRANSIT_PLANNED_ARRIVALS (без заголовків)")

    paths = planned_paths or _list_xlsx_files(planned_dir)
    if not paths:
        print(f"{_fmt_status('WRN', use_color)} | step=5f | reason=missing_xlsx_files | dir={os.path.abspath(planned_dir)}")
        return results_xlsx_path

    wb = openpyxl.load_workbook(results_xlsx_path)
    if SHEET_RINGS_SUMMARY not in wb.sheetnames:
        raise ValueError(f"results workbook: немає '{SHEET_RINGS_SUMMARY}'. Спочатку запусти build_rings_summary_sheet().")
    ws_sum = wb[SHEET_RINGS_SUMMARY]

    sym_to_row, sym_norm_to_row = _build_summary_row_index(ws_sum)
    mapping = _load_mapping(sku_master_path)
    group2_to_best = _load_group2_to_best_sku(sku_master_path)

    uz_totals = _load_uz_normalized_totals(wb)
    w39 = uz_totals.get("PR-DAE-39-1402-000-SET", 0.0)
    w49 = uz_totals.get("PR-DAE-49-1402-000-SET", 0.0)

    used_headers = { _norm_cell(ws_sum.cell(1,c).value) for c in range(1, ws_sum.max_column+1) if _norm_cell(ws_sum.cell(1,c).value) }

    for file_path in paths:
        file_base = os.path.splitext(os.path.basename(file_path))[0]
        awb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in awb.sheetnames:
            ws_a = awb[sheet_name]
            try:
                g2_min = _read_headerless_group2_qty_min(ws_a)
            except Exception as exc:
                print(f"{_fmt_status('WRN', use_color)} | step=5f | file={file_base} | sheet={sheet_name} | skipped | {exc}")
                continue

            # уникаємо колізій назв колонок між файлами/листами
            header = sheet_name
            if header in used_headers:
                header = f"{sheet_name} ({file_base})"
            used_headers.add(header)

            target_col = _ensure_column(ws_sum, header)

            filled = 0
            mapped_cnt = 0
            missing = 0
            split = 0

            sum_min_units = float(sum(g2_min.values()))
            sum_results_units = 0.0

            for g2, mn in g2_min.items():
                if g2 == "19-1402-000":
                    denom = (w39 * 3 + w49 * 4)
                    if denom <= 0:
                        missing += 1
                        continue
                    q39 = mn * (w39 / denom)
                    q49 = mn * (w49 / denom)
                    for sku, qty_units in [("PR-DAE-39-1402-000-SET", q39), ("PR-DAE-49-1402-000-SET", q49)]:
                        rr = sym_to_row.get(sku) or sym_norm_to_row.get(_normalize_sku(sku))
                        if rr is None:
                            missing += 1
                            continue
                        ws_sum.cell(rr, target_col, qty_units)
                        sum_results_units += qty_units
                        filled += 1
                    split += 1
                    continue

                sku_asm = group2_to_best.get(g2)
                if not sku_asm:
                    missing += 1
                    continue
                sku, asm = sku_asm
                qty_sets = mn / asm if asm else mn

                rr = sym_to_row.get(sku) or sym_norm_to_row.get(_normalize_sku(sku))
                if rr is None:
                    mapped_sku = mapping.get(_normalize_sku(sku))
                    if mapped_sku:
                        rr = sym_norm_to_row.get(mapped_sku)
                        if rr is not None:
                            mapped_cnt += 1

                if rr is None:
                    missing += 1
                    continue

                ws_sum.cell(rr, target_col, qty_sets)
                sum_results_units += qty_sets * asm
                filled += 1

            diff_units = abs(sum_min_units - sum_results_units)
            status = "OK" if diff_units <= TOL else "WRN"

            print(
                f"{_fmt_status(status, use_color)} | step=5f | file={file_base:<18} | sheet={sheet_name:<12} | "
                f"groups={len(g2_min):<4} | filled={filled:<4} | mapped={mapped_cnt:<3} | missing={missing:<3} | split={split:<2} | "
                f"sum_min_units={_fmt_num(sum_min_units):>8} | sum_results_units={_fmt_num(sum_results_units):>8} | diff_units={_fmt_num(diff_units):>8}"
            )

    # фінальне форматування
    _rotate_headers_vertical(ws_sum, start_col=5)
    _center_columns(ws_sum, cols=[1, 2, 3, 4], start_row=1)

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)} | step=5f | saved={os.path.abspath(results_xlsx_path)}")
    return results_xlsx_path


def _read_surowce_group2_min(ws) -> Dict[str, float]:
    """
    Повертає словник {Group2: MIN(B)} для листа surowce.
    Правило визначення Group2:
    - якщо A починається на 'PR-' => Group2 = A[7:18] (ПСТР(A;8;11))
    - інакше Group2 = A (вже спрощений)
    """
    col_a = _find_header_col(ws, "Номер по каталогу")
    col_b = _find_header_col(ws, "В С Е Г О")
    if col_a is None or col_b is None:
        raise ValueError(f"Лист '{ws.title}': не знайдено заголовки 'Номер по каталогу'/'В С Е Г О'")

    out: Dict[str, float] = {}
    rows_seen = 0

    for r in range(2, ws.max_row + 1):
        a = _norm_cell(ws.cell(r, col_a).value)
        if not a:
            continue
        b = _to_float(ws.cell(r, col_b).value)

        g2 = (a[7:18] if a.startswith("PR-") and len(a) >= 18 else a)

        if not g2:
            continue
        rows_seen += 1
        if g2 not in out:
            out[g2] = b
        else:
            out[g2] = min(out[g2], b)

    if rows_seen == 0:
        raise ValueError(f"Лист '{ws.title}': немає даних")
    return out


def _load_uz_normalized_totals(results_wb) -> Dict[str, float]:
    """
    Повертає dict: Symbol -> TOTAL з листа UZ-normalized.
    """
    if SHEET_UZ_NORM not in results_wb.sheetnames:
        return {}
    ws = results_wb[SHEET_UZ_NORM]
    col_sym = _find_header_col_any(ws, ["Symbol", "SKU", "Артикул", "SKU (ключ, унікальний)"])
    col_tot = _find_header_col_any(ws, ["TOTAL", "Total", "В С Е Г О", "ВСЕГО"])
    if col_sym is None or col_tot is None:
        return {}
    out: Dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        sym = _norm_cell(ws.cell(r, col_sym).value)
        if not sym:
            continue
        out[sym] = _to_float(ws.cell(r, col_tot).value)
    return out


def apply_stocks_minus_column_from_results(
    wb: "openpyxl.Workbook",
    ws_sum,
    sym_to_row: Dict[str, int],
    sym_norm_to_row: Dict[str, int],
    *,
    use_color: bool = True,
) -> None:
    """
    STEP 5e: якщо у results.xlsx є лист 'STOCKS-',
    заповнюємо колонку 'STOCKS-' у RINGS_SUMMARY значеннями -Quantity
    (тобто беремо з протилежним знаком), де Quantity — колонка C у листі STOCKS-.
    Матчинг по Symbol/SKU.
    """
    if "STOCKS-" not in wb.sheetnames:
        return

    ws_st = wb["STOCKS-"]

    # Symbol column: try header, else A
    col_sym = _find_header_col_any(ws_st, ["Symbol", "SKU", "Артикул", "SKU (ключ, унікальний)"])
    if col_sym is None:
        col_sym = 1

    # Quantity is column C by requirement; still allow header fallback if present
    col_qty = _find_header_col_any(ws_st, ["Quantity", "Qty", "Кількість", "Количество"])
    if col_qty is None:
        col_qty = 3

    mp: Dict[str, float] = {}
    sum_qty_selected = 0.0
    for r in range(2, ws_st.max_row + 1):
        sym = _norm_cell(ws_st.cell(r, col_sym).value)
        if not sym or not sym.startswith("PR-"):
            continue
        q = _to_float(ws_st.cell(r, col_qty).value)
        sum_qty_selected += q
        mp[sym] = mp.get(sym, 0.0) + (-q)

    col_target = _ensure_column(ws_sum, "STOCKS-")

    filled = 0
    missing = 0
    # очікувана сума = -сума Quantity лише для SKU, що починаються на 'PR-'
    sum_src = float(sum_qty_selected)
    sum_written = 0.0

    for sym, qneg in mp.items():
        rr = sym_to_row.get(sym) or sym_norm_to_row.get(_normalize_sku(sym))
        if rr is None:
            missing += 1
            continue
        ws_sum.cell(rr, col_target, qneg)
        sum_written += qneg
        filled += 1

    # Очікувана сума у колонці = -sum(original Quantity)
    expected = -sum_src
    diff = abs(expected - sum_written)
    status = "OK" if diff <= TOL else "WRN"

    print(
        f"{_fmt_status(status, use_color)} | step=5e | sheet=STOCKS- (results) | "
        f"in={len(mp):<6} | filled={filled:<6} | missing={missing:<6} | "
        f"sum_expected={_fmt_num(expected):>8} | sum_written={_fmt_num(sum_written):>8} | diff={_fmt_num(diff):>8}"
    )

def apply_surowce_workbook(
    results_xlsx_path: str,
    *,
    transit_dir: str = DEFAULT_TRANSIT_STOCK_STATUS_DIR,
    surowce_path: Optional[str] = None,
    sku_master_path: str = DEFAULT_SKU_MASTER_XLSX,
    use_color: bool = True,
) -> str:
    _print_banner("STEP 5d | Заповнення з surowce + перевірка в units")

    surowce_file = surowce_path or _find_first_xlsx(transit_dir, "surowce")
    if not surowce_file:
        print(f"{_fmt_status('WRN', use_color)} | step=5d | reason=missing_surowce_file | dir={os.path.abspath(transit_dir)}")
        return results_xlsx_path

    wb = openpyxl.load_workbook(results_xlsx_path)
    if SHEET_RINGS_SUMMARY not in wb.sheetnames:
        raise ValueError(f"results workbook: немає '{SHEET_RINGS_SUMMARY}'. Спочатку запусти build_rings_summary_sheet().")
    ws_sum = wb[SHEET_RINGS_SUMMARY]

    sym_to_row, sym_norm_to_row = _build_summary_row_index(ws_sum)
    mapping = _load_mapping(sku_master_path)
    group2_to_best = _load_group2_to_best_sku(sku_master_path)

    uz_totals = _load_uz_normalized_totals(wb)
    w39 = uz_totals.get("PR-DAE-39-1402-000-SET", 0.0)
    w49 = uz_totals.get("PR-DAE-49-1402-000-SET", 0.0)

    surowce_wb = openpyxl.load_workbook(surowce_file, data_only=True)

    for sheet_name in surowce_wb.sheetnames:
        ws_s = surowce_wb[sheet_name]
        try:
            g2_min = _read_surowce_group2_min(ws_s)
        except Exception as exc:
            print(f"{_fmt_status('WRN', use_color)} | step=5d | sheet={sheet_name} | skipped | {exc}")
            continue

        target_col = _ensure_column(ws_sum, sheet_name)

        filled = 0
        mapped_cnt = 0
        missing = 0
        split = 0

        sum_min_units = float(sum(g2_min.values()))
        sum_results_units = 0.0  # сума у "units" (повертаємо назад через AssemblyQty)

        # print(g2_min)
        for g2, mn in g2_min.items():
            # special split
            if g2 == "19-1402-000":
                denom = (w39 * 3 + w49 * 4)
                if denom <= 0:
                    missing += 1
                    continue
                q39 = mn * (w39 * 3 / denom)
                q49 = mn * (w49 * 4 / denom)
                for sku, qty_units in [("PR-DAE-39-1402-000-SET", q39), ("PR-DAE-49-1402-000-SET", q49)]:
                    rr = sym_to_row.get(sku) or sym_norm_to_row.get(_normalize_sku(sku))
                    if rr is None:
                        missing += 1
                        continue
                    ws_sum.cell(rr, target_col, qty_units)
                    sum_results_units += qty_units
                    filled += 1
                split += 1
                continue

            sku_asm = group2_to_best.get(g2)
            if not sku_asm:
                missing += 1
                continue
            sku, asm = sku_asm
            qty_sets = mn / asm if asm else mn

            rr = sym_to_row.get(sku) or sym_norm_to_row.get(_normalize_sku(sku))
            if rr is None:
                mapped_sku = mapping.get(_normalize_sku(sku))
                if mapped_sku:
                    rr = sym_norm_to_row.get(mapped_sku)
                    if rr is not None:
                        mapped_cnt += 1

            if rr is None:
                missing += 1
                continue

            ws_sum.cell(rr, target_col, qty_sets)
            sum_results_units += qty_sets * asm
            filled += 1

        diff_units = abs(sum_min_units - sum_results_units)
        status = "OK" if diff_units <= TOL else "WRN"

        print(
            f"{_fmt_status(status, use_color)} | step=5d | sheet={sheet_name:<20} | "
            f"groups={len(g2_min):<4} | filled={filled:<4} | mapped={mapped_cnt:<3} | missing={missing:<3} | split={split:<2} | "
            f"sum_min_units={_fmt_num(sum_min_units):>8} | sum_results_units={_fmt_num(sum_results_units):>8} | diff_units={_fmt_num(diff_units):>8}"
        )

    # STEP 5e: колонка STOCKS- з results.xlsx (мінус Quantity)
    apply_stocks_minus_column_from_results(wb, ws_sum, sym_to_row, sym_norm_to_row, use_color=use_color)
    # Порядок: STOCKS- одразу після Monika
    _move_column_after(ws_sum, col_name="STOCKS-", after_col_name="Monika")

    # вертикальні заголовки
    _rotate_headers_vertical(ws_sum, start_col=5)
    # центруємо базові колонки
    _center_columns(ws_sum, cols=[1, 2, 3, 4], start_row=1)

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)} | step=5d | saved={os.path.abspath(results_xlsx_path)}")
    return results_xlsx_path


def generate_rings_summary(
    results_xlsx_path: str = DEFAULT_RESULTS_XLSX,
    *,
    transit_stock_status_dir: str = DEFAULT_TRANSIT_STOCK_STATUS_DIR,
    transit_planned_arrivals_dir: str = DEFAULT_TRANSIT_PLANNED_ARRIVALS_DIR,
    sets_path: Optional[str] = None,
    surowce_path: Optional[str] = None,
    planned_paths: Optional[List[str]] = None,
    sku_master_path: str = DEFAULT_SKU_MASTER_XLSX,
    use_color: bool = True,
) -> str:
    """Один виклик = весь STEP 5 (5a→5f)."""
    results = build_rings_summary_sheet(
        results_xlsx_path=results_xlsx_path,
        sku_master_path=sku_master_path,
        use_color=use_color,
    )
    results = apply_transit_stock_sets(
        results_xlsx_path=results,
        transit_dir=transit_stock_status_dir,
        sets_path=sets_path,
        surowce_path=surowce_path,
        sku_master_path=sku_master_path,
        apply_surowce=True,
        use_color=use_color,
    )
    # planned arrivals
    results = apply_planned_arrivals_workbooks(
        results_xlsx_path=results,
        planned_dir=transit_planned_arrivals_dir,
        planned_paths=planned_paths,
        sku_master_path=sku_master_path,
        use_color=use_color,
    )
    # STEP 5g: Order_1..3 + Total Available (формули)
    wb = openpyxl.load_workbook(results)
    if SHEET_RINGS_SUMMARY not in wb.sheetnames:
        raise ValueError("results workbook: немає листа 'RINGS_SUMMARY' для STEP 5g")
    ws_sum = wb[SHEET_RINGS_SUMMARY]
    add_orders_and_total_available(ws_sum)
    _rotate_headers_vertical(ws_sum, start_col=5)
    _center_columns(ws_sum, cols=[1, 2, 3, 4], start_row=1)
    wb.save(results)
    print(f"{_fmt_status('OK', use_color)} | step=5g | added=Order_1,Order_2,Order_3,Total Available | saved={os.path.abspath(results)}")
    return results


if __name__ == "__main__":
    generate_rings_summary()