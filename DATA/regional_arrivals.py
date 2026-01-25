"""
step2_regional_arrivals.py

КРОК 2 пайплайна:

Папка REGIONAL_ARRIVALS містить 0..3 Excel-файли "надходжень" по регіонах (UA/KZ/UZ).
У файлах можливі дублікати артикулів — потрібно сумувати кількість.

Результат: у книзі SALES/RESULT/results.xlsx створити/замінити 3 листи:
  UA+, KZ+, UZ+
і розмістити їх ПЕРЕД 6-ма листами, що вже є наприкінці книги:
  UA-normalized, KZ-normalized, UZ-normalized, UA-sales, KZ-sales, UZ-sales

Формат листів UA+/KZ+/UZ+:
  1) Symbol_raw          (як у файлі надходжень, з колонки "№ по каталогу"/"Артикул"/...)
  2) Symbol              (унікальний, після normalize + mapping)
     - якщо SKU є у sku_master -> без заливки
     - якщо виправили mapping-ом -> зелена заливка
     - якщо не знайдено навіть після mapping -> червона заливка
  3) Quantity            (сума по Symbol)

Нормалізація/валідація як у кроці 3:
- прибираємо всі пробіли (в т.ч. NBSP)
- заміна схожих кириличних символів на латиницю
- перевірка в sku_master.xlsx (лист SKU, колонка "SKU (ключ, унікальний)")
- mapping (лист Mapping: "Неправильный" -> "Артикул")

Консоль:
- банер
- 1 рядок на регіон (OK/WRN/ERR) у стилі кроків 1, 3
- фінальний рядок: OK | step=2 | saved=...
"""

from __future__ import annotations

from typing import Dict, List, Optional, Set, Tuple

import datetime
import os
from pathlib import Path
import re

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Стилі та загальні хелпери (як у Step 2)
# =============================================================================
FILL_RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
FILL_GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
FILL_CLEAR = PatternFill()

_WS_RE = re.compile(r"[\s\u00A0\u2007\u202F]+")
_C2L = {
    "А": "A", "В": "B", "Е": "E", "І": "I", "К": "K", "М": "M",
    "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
    "а": "A", "в": "B", "е": "E", "і": "I", "к": "K", "м": "M",
    "н": "H", "о": "O", "р": "P", "с": "C", "т": "T", "х": "X",
    "Ё": "E", "ё": "E",
    "Й": "I", "й": "I",
    "З": "3", "з": "3",
    "У": "Y", "у": "Y",
}


def _print_banner(title: str) -> None:
    line = "=" * 72
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(line)
    print(f"{title} | {ts}")
    print(line)


def _fmt_status(tag: str, use_color: bool = True) -> str:
    """OK/WRN/ERR з опціональним ANSI-кольором."""
    if not use_color:
        return tag
    colors = {
        "OK": "\x1b[32m",
        "WRN": "\x1b[33m",
        "ERR": "\x1b[31m",
    }
    reset = "\x1b[0m"
    return f"{colors.get(tag, '')}{tag}{reset}"


def normalize_symbol(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = _WS_RE.sub("", s)
    s = "".join(_C2L.get(ch, ch) for ch in s)
    return s


def _autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 5000) -> None:
    try:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, min(ws.max_row, max_rows) + 1):
                v = ws.cell(r, col_idx).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)
    except Exception:
        pass


# =============================================================================
# SKU master
# =============================================================================
# =============================================================================
# Шляхи проєкту (прив'язані до розташування цього файлу)
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent          # .../DATA
SKU_DIR = BASE_DIR / "SKU"
SALES_DIR = BASE_DIR / "SALES"

REGIONAL_ARRIVALS_DIR = SALES_DIR / "REGIONAL_ARRIVALS"
RESULT_DIR = SALES_DIR / "RESULT"
RESULTS_XLSX = RESULT_DIR / "results.xlsx"

# Дефолтний довідник SKU
DEFAULT_SKU_MASTER_PATH = str(SKU_DIR / "sku_master.xlsx")

# Створюємо каталоги, якщо їх ще немає
RESULT_DIR.mkdir(parents=True, exist_ok=True)

SKU_SHEET_NAME = "SKU"
SKU_HEADER_NAME = "SKU (ключ, унікальний)"
MAPPING_SHEET_NAME = "Mapping"
MAPPING_WRONG_HEADER = "Неправильный"
MAPPING_RIGHT_HEADER = "Артикул"


def _find_header_col(ws: openpyxl.worksheet.worksheet.Worksheet, header_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is not None and str(v).strip() == header_name:
            return col
    return None


def load_sku_set_and_mapping(sku_master_path: str) -> Tuple[Set[str], Dict[str, str]]:
    wb = openpyxl.load_workbook(sku_master_path, data_only=True)

    if SKU_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"У sku_master.xlsx немає листа '{SKU_SHEET_NAME}'")

    ws_sku = wb[SKU_SHEET_NAME]
    sku_col = _find_header_col(ws_sku, SKU_HEADER_NAME)
    if sku_col is None:
        raise ValueError(
            f"На листі '{SKU_SHEET_NAME}' не знайдено заголовок '{SKU_HEADER_NAME}' у 1-му рядку"
        )

    sku_set: Set[str] = set()
    for r in range(2, ws_sku.max_row + 1):
        v = normalize_symbol(ws_sku.cell(r, sku_col).value)
        if v:
            sku_set.add(v)

    mapping: Dict[str, str] = {}
    if MAPPING_SHEET_NAME in wb.sheetnames:
        ws_map = wb[MAPPING_SHEET_NAME]
        wrong_col = _find_header_col(ws_map, MAPPING_WRONG_HEADER)
        right_col = _find_header_col(ws_map, MAPPING_RIGHT_HEADER)
        if wrong_col is None or right_col is None:
            raise ValueError(
                f"На листі '{MAPPING_SHEET_NAME}' очікуються заголовки "
                f"'{MAPPING_WRONG_HEADER}' та '{MAPPING_RIGHT_HEADER}' у 1-му рядку"
            )

        for r in range(2, ws_map.max_row + 1):
            wrong = normalize_symbol(ws_map.cell(r, wrong_col).value)
            right = normalize_symbol(ws_map.cell(r, right_col).value)
            if wrong and right:
                mapping[wrong] = right

    return sku_set, mapping


# =============================================================================
# Парсинг надходжень (детект колонок)
# =============================================================================
_PUNCT_RE = re.compile(r"[^0-9a-zа-яіїєёґ\s]+", flags=re.IGNORECASE)

SKU_KEYS_RAW = [
    "№ по каталогу",
    "номер по каталогу",
    "по каталогу",
    "каталог",
    "каталожный номер",
    "кат номер",
    "кат. номер",
    "артикул",
    "sku",
    "symbol",
    "item",
    "part number",
]
QTY_KEYS_RAW = [
    "кол-во",
    "кол во",
    "количество",
    "qty",
    "quantity",
    "кол",
    "count",
    "pcs",
    "заказ",
    "замовлення",
    "кількість",
    "amount",
    "order"
]


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    t = str(value).replace("\u00A0", " ").strip().lower()
    t = _PUNCT_RE.sub(" ", t)
    t = " ".join(t.split())
    return t


def _build_keys(raw: List[str]) -> List[str]:
    uniq = {_norm_text(x) for x in raw if _norm_text(x)}
    return sorted(uniq, key=len, reverse=True)


SKU_KEYS = _build_keys(SKU_KEYS_RAW)
QTY_KEYS = _build_keys(QTY_KEYS_RAW)


def _contains_any(text: str, keys: List[str]) -> bool:
    return bool(text) and any(k in text for k in keys)


def detect_arrivals_header(ws: openpyxl.worksheet.worksheet.Worksheet, scan_rows: int = 50, scan_cols: int = 40) -> Tuple[int, int, int]:
    """Повертає (header_row, sku_col, qty_col) 1-based."""
    best = None  # (score, header_row, sku_col, qty_col)
    max_r = min(ws.max_row, scan_rows)
    max_c = min(ws.max_column, scan_cols)

    for r in range(1, max_r + 1):
        # collect normalized headers
        headers = {c: _norm_text(ws.cell(r, c).value) for c in range(1, max_c + 1)}
        sku_col = next((c for c, h in headers.items() if _contains_any(h, SKU_KEYS)), None)
        qty_col = next((c for c, h in headers.items() if _contains_any(h, QTY_KEYS)), None)

        score = 0
        score += 10 if sku_col else 0
        score += 10 if qty_col else 0
        if score >= 20:
            # prefer earlier header rows but stable
            score_adj = score - (r * 0.01)
            cand = (score_adj, r, sku_col or 1, qty_col or 1)
            if best is None or cand[0] > best[0]:
                best = cand

    if best is None:
        raise ValueError("Не знайдено шапку: не вдалося визначити колонки SKU/Quantity.")
    _, header_row, sku_col, qty_col = best
    return header_row, sku_col, qty_col


def detect_region_from_filename(filename: str) -> str:
    up = filename.upper()
    if "UA" in up or "ЮА" in up:
        return "UA"
    if "KZ" in up:
        return "KZ"
    if "UZ" in up:
        return "UZ"
    return "UNK"


def _to_float(x: object) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    try:
        s = str(x).strip().replace(",", ".")
        return float(s) if s else 0.0
    except Exception:
        return 0.0


# =============================================================================
# Головна утиліта Step 2
# =============================================================================
def build_regional_arrivals_sheets(
    results_xlsx_path: str,
    regional_arrivals_dir: Optional[str] = None,
    sku_master_path: Optional[str] = None,
    *,
    use_color: bool = True,
) -> str:
    """
    Читає 0..3 файли з REGIONAL_ARRIVALS і створює/замінює листи UA+, KZ+, UZ+ у results.xlsx.
    """
    _print_banner("STEP 2 | Build regional arrivals sheets")

    # У Colab/CLI зручно, коли дефолтні шляхи не залежать від поточної директорії
    if regional_arrivals_dir is None:
        regional_arrivals_dir = str(REGIONAL_ARRIVALS_DIR)


    # --- SKU master (optional) ---
    sku_master_path = sku_master_path or DEFAULT_SKU_MASTER_PATH
    sku_set: Set[str] = set()
    mapping: Dict[str, str] = {}
    try:
        if not os.path.exists(sku_master_path):
            raise FileNotFoundError(f"sku_master.xlsx не знайдено: {sku_master_path}")
        sku_set, mapping = load_sku_set_and_mapping(sku_master_path)
    except Exception as exc:
        print(f"{_fmt_status('WRN', use_color)}  | step=2 | sku_master_not_loaded | {exc}")

    if not os.path.exists(results_xlsx_path):
        raise FileNotFoundError(f"results.xlsx не знайдено: {results_xlsx_path}")

    wb = openpyxl.load_workbook(results_xlsx_path)

    # Collect arrivals files (xlsx only)
    arrivals_files: List[str] = []
    if os.path.isdir(regional_arrivals_dir):
        arrivals_files = [
            f for f in os.listdir(regional_arrivals_dir)
            if f.lower().endswith(".xlsx") and not f.startswith("~$")
        ]

    # Choose by region (first match). If multiple — pick lexicographically (stable)
    region_to_file: Dict[str, Optional[str]] = {"UA": None, "KZ": None, "UZ": None}
    for f in sorted(arrivals_files):
        reg = detect_region_from_filename(f)
        if reg in region_to_file and region_to_file[reg] is None:
            region_to_file[reg] = os.path.join(regional_arrivals_dir, f)

    # Create/replace three sheets
    for reg in ("UA", "KZ", "UZ"):
        sheet_title = f"{reg}+"
        if sheet_title in wb.sheetnames:
            wb.remove(wb[sheet_title])
        ws_out = wb.create_sheet(sheet_title)

        # header
        ws_out.cell(1, 1, "Symbol_raw")
        ws_out.cell(1, 2, "Symbol")
        ws_out.cell(1, 3, "Quantity")

        src_path = region_to_file.get(reg)
        if not src_path or not os.path.exists(src_path):
            _autosize_columns(ws_out)
            print(f"{_fmt_status('WRN', use_color)}  | region={reg:<2} | in={0:<5} | unique={0:<5} | merged={0:<5} | green={0:<4} | red={0:<4} | saved={sheet_title} | reason=missing_file")
            continue

        try:
            wb_src = openpyxl.load_workbook(src_path, data_only=True)
            ws_src = wb_src[wb_src.sheetnames[0]]

            header_row, sku_col, qty_col = detect_arrivals_header(ws_src)

            # aggregate by final symbol
            agg_qty: Dict[str, float] = {}
            agg_raw: Dict[str, str] = {}   # keep first raw value per final key
            agg_fill: Dict[str, Optional[PatternFill]] = {}

            in_rows = 0
            for r in range(header_row + 1, ws_src.max_row + 1):
                raw = ws_src.cell(r, sku_col).value
                qty = ws_src.cell(r, qty_col).value

                raw_s = "" if raw is None else str(raw).strip()
                if not raw_s:
                    continue

                in_rows += 1
                norm = normalize_symbol(raw_s)
                final = norm
                fill = None

                if final and sku_set:
                    if final in sku_set:
                        fill = None
                    else:
                        mapped = mapping.get(final)
                        if mapped and mapped in sku_set:
                            final = mapped
                            fill = FILL_GREEN
                        else:
                            fill = FILL_RED

                if final not in agg_qty:
                    agg_qty[final] = 0.0
                    agg_raw[final] = raw_s
                    agg_fill[final] = fill
                else:
                    # keep "worst" fill (red > green > none)
                    prev = agg_fill.get(final)
                    cur = fill
                    if prev == FILL_RED or cur == FILL_RED:
                        agg_fill[final] = FILL_RED
                    elif prev == FILL_GREEN or cur == FILL_GREEN:
                        agg_fill[final] = FILL_GREEN
                    else:
                        agg_fill[final] = None

                agg_qty[final] += _to_float(qty)

            # write
            out_r = 2
            for sym in sorted(agg_qty.keys()):
                ws_out.cell(out_r, 1, agg_raw.get(sym, sym))
                c2 = ws_out.cell(out_r, 2, sym)
                ws_out.cell(out_r, 3, agg_qty[sym])

                fill = agg_fill.get(sym)
                if fill is not None:
                    c2.fill = fill
                else:
                    c2.fill = FILL_CLEAR

                out_r += 1

            _autosize_columns(ws_out)

            unique = len(agg_qty)
            merged = max(0, in_rows - unique)
            green = sum(1 for f in agg_fill.values() if f == FILL_GREEN)
            red = sum(1 for f in agg_fill.values() if f == FILL_RED)

            print(
                f"{_fmt_status('OK', use_color)}   | region={reg:<2} | in={in_rows:<5} | unique={unique:<5} | merged={merged:<5} | green={green:<4} | red={red:<4} | saved={sheet_title} | file={os.path.basename(src_path)}"
            )

        except Exception as exc:
            # Keep empty sheet but report error
            print(f"{_fmt_status('ERR', use_color)}  | region={reg:<2} | saved={sheet_title} | error={exc}")

    # Reorder: UA+,KZ+,UZ+ before the 6 tail sheets
    arrivals_titles = ["UA+", "KZ+", "UZ+"]
    tail6 = ["UA-normalized", "KZ-normalized", "UZ-normalized", "UA-sales", "KZ-sales", "UZ-sales"]

    others = [ws for ws in wb._sheets if ws.title not in arrivals_titles + tail6]
    arrivals_sheets = [wb[t] for t in arrivals_titles if t in wb.sheetnames]
    tail_sheets = [wb[t] for t in tail6 if t in wb.sheetnames]
    wb._sheets = others + arrivals_sheets + tail_sheets

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)}   | step=2 | saved={os.path.abspath(results_xlsx_path)}")
    return os.path.abspath(results_xlsx_path)


def main() -> None:
    """CLI entrypoint.

    Типовий сценарій у Colab:
        !python DATA/regional_arrivals.py

    За потреби можна передати інші шляхи через аргументи.
    """
    import argparse

    parser = argparse.ArgumentParser(description="Крок 3: побудова аркушів регіональних надходжень (UA+/KZ+/UZ+).")
    parser.add_argument(
        "--results",
        default=str(RESULTS_XLSX),
        help="Шлях до results.xlsx (за замовчуванням: DATA/SALES/RESULT/results.xlsx).",
    )
    parser.add_argument(
        "--regional-arrivals-dir",
        default=str(REGIONAL_ARRIVALS_DIR),
        help="Папка з файлами REGIONAL_ARRIVALS (за замовчуванням: DATA/SALES/REGIONAL_ARRIVALS).",
    )
    parser.add_argument(
        "--sku-master",
        default=DEFAULT_SKU_MASTER_PATH,
        help="Шлях до sku_master.xlsx (за замовчуванням: DATA/SKU/sku_master.xlsx).",
    )
    parser.add_argument(
        "--no-color",
        action="store_true",
        help="Вимкнути кольорові статуси в консолі.",
    )

    args = parser.parse_args()

    build_regional_arrivals_sheets(
        results_xlsx_path=args.results,
        regional_arrivals_dir=args.regional_arrivals_dir,
        sku_master_path=args.sku_master,
        use_color=not args.no_color,
    )


if __name__ == "__main__":
    main()
