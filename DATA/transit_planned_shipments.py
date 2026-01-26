"""
transit_planned_shipments.py (production)

STEP 4: Планові відвантаження (TRANSIT_PLANNED_SHIPMENTS) -> лист "STOCKS-" у results.xlsx.

Що робимо:
- Читаємо 0..3 Excel-файли з папки SALES/TRANSIT_PLANNED_SHIPMENTS
- Агрегуємо кількість по SKU (дублі сумуються всередині і між файлами)
- Нормалізуємо SKU та валідуємо через SKU/Mapping у SKU/sku_master.xlsx
- Створюємо/перезаписуємо лист "STOCKS-" у SALES/RESULT/results.xlsx з колонками:
    Symbol_raw | Symbol | Quantity

Розфарбовування "Symbol":
- валідний SKU -> без заливки
- виправлено через Mapping -> зелений
- все ще невалідний -> червоний

Порядок листів: "STOCKS-" ставимо перед групою з 9 листів:
UA+, KZ+, UZ+, UA-normalized, KZ-normalized, UZ-normalized, UA-sales, KZ-sales, UZ-sales
"""


from __future__ import annotations

from typing import Dict, List, Optional, Set, Tuple

import datetime
import os
import re

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ------------------ console helpers ------------------
FILL_RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
FILL_GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
FILL_CLEAR = PatternFill()

_WS_RE = re.compile(r"[\s\u00A0\u2007\u202F]+")
_C2L = {
    "А": "A", "В": "B", "Е": "E", "І": "I", "К": "K", "М": "M",
    "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
    "а": "A", "в": "B", "е": "E", "і": "I", "к": "K", "м": "M",
    "н": "H", "о": "O", "р": "P", "с": "C", "т": "T", "х": "X",
    "Ё": "E", "ё": "E",
    "Й": "I", "й": "I",
    "З": "3", "з": "3",
    "У": "Y", "у": "Y",
}

_PUNCT_RE = re.compile(r"[^0-9a-zа-яіїєёґ\s]+", flags=re.IGNORECASE)


def _print_banner(title: str) -> None:
    line = "=" * 72
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(line)
    print(f"{title} | {ts}")
    print(line)


def _fmt_status(tag: str, use_color: bool = True) -> str:
    if not use_color:
        return tag
    colors = {"OK": "\x1b[32m", "WRN": "\x1b[33m", "ERR": "\x1b[31m"}
    reset = "\x1b[0m"
    return f"{colors.get(tag, '')}{tag}{reset}"


def _autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 5000) -> None:
    try:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, min(ws.max_row, max_rows) + 1):
                v = ws.cell(r, col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)
    except Exception:
        pass


def normalize_symbol(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = _WS_RE.sub("", s)
    s = "".join(_C2L.get(ch, ch) for ch in s)
    return s


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    t = str(value).replace("\u00A0", " ").strip().lower()
    t = _PUNCT_RE.sub(" ", t)
    return " ".join(t.split())


def _contains_any(text: str, keys: List[str]) -> bool:
    return bool(text) and any(k in text for k in keys)


def _build_keys(raw: List[str]) -> List[str]:
    uniq = {_norm_text(x) for x in raw if _norm_text(x)}
    return sorted(uniq, key=len, reverse=True)


def _to_float(x: object) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    try:
        s = str(x).strip().replace(",", ".")
        return float(s) if s else 0.0
    except Exception:
        return 0.0


# ------------------ шляхи проєкту ------------------
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
SALES_ROOT_DIR = BASE_DIR / "SALES"

# Вхідні папки/файли
TRANSIT_PLANNED_SHIPMENTS_DIR = SALES_ROOT_DIR / "TRANSIT_PLANNED_SHIPMENTS"
RESULT_DIR = SALES_ROOT_DIR / "RESULT"
RESULTS_XLSX = RESULT_DIR / "results.xlsx"

SKU_MASTER_DIR = BASE_DIR / "SKU"
DEFAULT_SKU_MASTER_PATH = str(SKU_MASTER_DIR / "sku_master.xlsx")


SKU_SHEET_NAME = "SKU"
SKU_HEADER_NAME = "SKU (ключ, унікальний)"
MAPPING_SHEET_NAME = "Mapping"
MAPPING_WRONG_HEADER = "Неправильный"
MAPPING_RIGHT_HEADER = "Артикул"


def _find_header_col(ws: openpyxl.worksheet.worksheet.Worksheet, header_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is not None and str(v).strip() == header_name:
            return col
    return None


def load_sku_set_and_mapping(sku_master_path: str) -> Tuple[Set[str], Dict[str, str]]:
    wb = openpyxl.load_workbook(sku_master_path, data_only=True)

    if SKU_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"У sku_master.xlsx немає листа '{SKU_SHEET_NAME}'")

    ws_sku = wb[SKU_SHEET_NAME]
    sku_col = _find_header_col(ws_sku, SKU_HEADER_NAME)
    if sku_col is None:
        raise ValueError(f"Не знайдено '{SKU_HEADER_NAME}' у листі '{SKU_SHEET_NAME}'")

    sku_set: Set[str] = set()
    for r in range(2, ws_sku.max_row + 1):
        sku = normalize_symbol(ws_sku.cell(r, sku_col).value)
        if sku:
            sku_set.add(sku)

    mapping: Dict[str, str] = {}
    if MAPPING_SHEET_NAME in wb.sheetnames:
        ws_map = wb[MAPPING_SHEET_NAME]
        wrong_col = _find_header_col(ws_map, MAPPING_WRONG_HEADER)
        right_col = _find_header_col(ws_map, MAPPING_RIGHT_HEADER)
        if wrong_col is None or right_col is None:
            raise ValueError("Не знайдено заголовки Mapping у sku_master.xlsx")
        for r in range(2, ws_map.max_row + 1):
            wrong = normalize_symbol(ws_map.cell(r, wrong_col).value)
            right = normalize_symbol(ws_map.cell(r, right_col).value)
            if wrong and right:
                mapping[wrong] = right

    return sku_set, mapping


# ------------------ header detection ------------------
SKU_KEYS = _build_keys([
    "№ по каталогу", "номер по каталогу", "по каталогу", "каталог",
    "каталожный номер", "кат номер", "кат. номер",
    "артикул", "sku", "symbol", "item", "part number",
])

QTY_KEYS = _build_keys([
    "кол-во", "кол во", "количество", "qty", "quantity", "кол", "count", "pcs",
    "заказ",
])


def detect_header(ws: openpyxl.worksheet.worksheet.Worksheet, scan_rows: int = 50, scan_cols: int = 40) -> Tuple[int, int, int]:
    best = None
    max_r = min(ws.max_row, scan_rows)
    max_c = min(ws.max_column, scan_cols)

    for r in range(1, max_r + 1):
        headers = {c: _norm_text(ws.cell(r, c).value) for c in range(1, max_c + 1)}
        sku_col = next((c for c, h in headers.items() if _contains_any(h, SKU_KEYS)), None)
        qty_col = next((c for c, h in headers.items() if _contains_any(h, QTY_KEYS)), None)

        score = (10 if sku_col else 0) + (10 if qty_col else 0)
        if score >= 20:
            score_adj = score - (r * 0.01)
            cand = (score_adj, r, sku_col or 1, qty_col or 1)
            if best is None or cand[0] > best[0]:
                best = cand

    if best is None:
        raise ValueError("Не знайдено шапку: не вдалося визначити колонки SKU/Quantity.")
    _, header_row, sku_col, qty_col = best
    return header_row, sku_col, qty_col


# ------------------ main step 4 ------------------
def build_planned_shipments_sheet(
    results_xlsx_path: str,
    transit_planned_shipments_dir: str = str(TRANSIT_PLANNED_SHIPMENTS_DIR),
    sku_master_path: Optional[str] = None,
    *,
    use_color: bool = True,
) -> str:
    _print_banner("STEP 4 | Build planned shipments sheet (STOCKS-)")

    sku_master_path = sku_master_path or DEFAULT_SKU_MASTER_PATH
    sku_set: Set[str] = set()
    mapping: Dict[str, str] = {}
    try:
        if not os.path.exists(sku_master_path):
            raise FileNotFoundError(f"sku_master.xlsx не знайдено: {sku_master_path}")
        sku_set, mapping = load_sku_set_and_mapping(sku_master_path)
    except Exception as exc:
        print(f"{_fmt_status('WRN', use_color)}  | step=4 | sku_master_not_loaded | {exc}")

    if not os.path.exists(results_xlsx_path):
        raise FileNotFoundError(f"results.xlsx не знайдено: {results_xlsx_path}")

    wb = openpyxl.load_workbook(results_xlsx_path)

    sheet_title = "STOCKS-"
    if sheet_title in wb.sheetnames:
        wb.remove(wb[sheet_title])
    ws_out = wb.create_sheet(sheet_title)

    ws_out.cell(1, 1, "Symbol_raw")
    ws_out.cell(1, 2, "Symbol")
    ws_out.cell(1, 3, "Quantity")

    files: List[str] = []
    if os.path.isdir(transit_planned_shipments_dir):
        files = [f for f in os.listdir(transit_planned_shipments_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    files = sorted(files)

    agg_qty: Dict[str, float] = {}
    agg_raw: Dict[str, str] = {}
    agg_fill: Dict[str, Optional[PatternFill]] = {}

    if not files:
        _autosize_columns(ws_out)
        print(f"{_fmt_status('WRN', use_color)}  | file=-{'':<28} | in={0:<5} | unique={0:<5} | merged={0:<5} | green={0:<4} | red={0:<4} | saved={sheet_title} | reason=empty_folder")
    else:
        for f in files:
            src_path = os.path.join(transit_planned_shipments_dir, f)
            try:
                wb_src = openpyxl.load_workbook(src_path, data_only=True)
                ws_src = wb_src[wb_src.sheetnames[0]]
                header_row, sku_col, qty_col = detect_header(ws_src)

                in_rows = 0
                per_file: Dict[str, int] = {}

                for r in range(header_row + 1, ws_src.max_row + 1):
                    raw = ws_src.cell(r, sku_col).value
                    qty = ws_src.cell(r, qty_col).value
                    raw_s = "" if raw is None else str(raw).strip()
                    if not raw_s:
                        continue

                    in_rows += 1
                    norm = normalize_symbol(raw_s)
                    final = norm
                    fill = None

                    if final and sku_set:
                        if final in sku_set:
                            fill = None
                        else:
                            mapped = mapping.get(final)
                            if mapped and mapped in sku_set:
                                final = mapped
                                fill = FILL_GREEN
                            else:
                                fill = FILL_RED

                    per_file[final] = per_file.get(final, 0) + 1

                    if final not in agg_qty:
                        agg_qty[final] = 0.0
                        agg_raw[final] = raw_s
                        agg_fill[final] = fill
                    else:
                        prev = agg_fill.get(final)
                        cur = fill
                        if prev == FILL_RED or cur == FILL_RED:
                            agg_fill[final] = FILL_RED
                        elif prev == FILL_GREEN or cur == FILL_GREEN:
                            agg_fill[final] = FILL_GREEN
                        else:
                            agg_fill[final] = None

                    agg_qty[final] += _to_float(qty)

                unique_in_file = len(per_file)
                merged_in_file = max(0, in_rows - unique_in_file)
                green_in_file = sum(1 for k in per_file if agg_fill.get(k) == FILL_GREEN)
                red_in_file = sum(1 for k in per_file if agg_fill.get(k) == FILL_RED)

                print(
                    f"{_fmt_status('OK', use_color)}   | file={f[:30]:<30} | in={in_rows:<5} | unique={unique_in_file:<5} | merged={merged_in_file:<5} | green={green_in_file:<4} | red={red_in_file:<4} | saved={sheet_title}"
                )

            except Exception as exc:
                print(f"{_fmt_status('ERR', use_color)}  | file={f[:30]:<30} | saved={sheet_title} | error={exc}")

        # write aggregated
        out_r = 2
        for sym in sorted(agg_qty.keys()):
            ws_out.cell(out_r, 1, agg_raw.get(sym, sym))
            c2 = ws_out.cell(out_r, 2, sym)
            ws_out.cell(out_r, 3, agg_qty[sym])
            fill = agg_fill.get(sym)
            c2.fill = fill if fill is not None else FILL_CLEAR
            out_r += 1
        _autosize_columns(ws_out)

    # reorder: STOCKS- before UA+,KZ+,UZ+ and the 6 tail sheets
    arrivals3 = ["UA+", "KZ+", "UZ+"]
    tail6 = ["UA-normalized", "KZ-normalized", "UZ-normalized", "UA-sales", "KZ-sales", "UZ-sales"]
    group9 = [sheet_title] + arrivals3 + tail6
    others = [ws for ws in wb._sheets if ws.title not in group9]
    group_sheets = [wb[t] for t in group9 if t in wb.sheetnames]
    wb._sheets = others + group_sheets

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)}   | step=4 | saved={os.path.abspath(results_xlsx_path)}")
    return os.path.abspath(results_xlsx_path)


def main() -> None:
    build_planned_shipments_sheet(str(RESULTS_XLSX))


if __name__ == "__main__":
    main()
