"""
normalize_results_workbook.py

КРОК 3 (після build_results_workbook + build_regional_arrivals_sheets):

- Читає листи: UA-sales, KZ-sales, UZ-sales з книги results.xlsx
- Створює/замінює листи: UA-normalized, KZ-normalized, UZ-normalized у ТІЙ ЖЕ книзі

Логіка *-normalized:
1) Беремо Symbol -> нормалізуємо (прибрати всі пробіли + кирилиця->латиниця) => Symbol_normalized
2) Валідація по SKU master (SKU/sku_master.xlsx):
   - якщо не знайдено -> пробуємо Mapping (зелений) або червоний
3) У *-normalized Symbol_normalized має бути УНІКАЛЬНИМ:
   - якщо виникають дублікати (після мапінгу) — агрегуємо рядки по Symbol_normalized
   - числові колонки (місяці, Stock, TOTAL) сумуються
4) Після Symbol_normalized додаємо 2 колонки:
   - Mediana: медіана місячних продажів (медіана по колонках місяців)
   - Month_sales: = Mediana, заливка жовта (щоб можна було правити руками)
5) Stock переноситься (виводиться) одразу після Month_sales
6) Порядок листів у кінці книги:
   UA-normalized, KZ-normalized, UZ-normalized, UA-sales, KZ-sales, UZ-sales

Консоль:
- Банер старту
- WARNING тільки якщо є проблеми (нема SKU master / нема листа)
- Saved: <path>
"""

from __future__ import annotations

from typing import Dict, List, Optional, Set, Tuple

import datetime
import os
from pathlib import Path
import re
import statistics

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Константи/налаштування
# =============================================================================
SYMBOL_HEADER_NAME = "Symbol"
NORMALIZED_COL_HEADER = "Symbol_normalized"


# =============================================================================
# Уніфікація назв місяців (вихідні колонки завжди: янв..дек)
# =============================================================================
REQUIRED_COLUMNS = [
    "Symbol_normalized",
    "янв", "фев", "мар", "апр", "май", "июн",
    "июл", "авг", "сен", "окт", "ноя", "дек",
]

# Канонічні місяці -> синоніми, які можуть зустрічатися у заголовках
_MONTH_ALIASES = {
    "янв": ["янв", "январ", "january", "jan"],
    "фев": ["фев", "феврал", "february", "feb"],
    "мар": ["мар", "март", "march", "mar"],
    "апр": ["апр", "апрел", "april", "apr"],
    "май": ["май", "мая", "may"],
    "июн": ["июн", "июнь", "june", "jun"],
    "июл": ["июл", "июль", "july", "jul"],
    "авг": ["авг", "август", "august", "aug"],
    "сен": ["сен", "сентябр", "september", "sep", "sept"],
    "окт": ["окт", "октябр", "october", "oct"],
    "ноя": ["ноя", "ноябр", "november", "nov"],
    "дек": ["дек", "декабр", "december", "dec"],
}

def _canon_month(header: str) -> Optional[str]:
    """Повертає 'янв'..'дек', якщо заголовок схожий на назву місяця."""
    if header is None:
        return None
    h = str(header).strip().lower()
    if not h:
        return None
    for m, keys in _MONTH_ALIASES.items():
        if any(k in h for k in keys):
            return m
    return None

# =============================================================================
# Шляхи проєкту (прив'язані до розташування цього файлу)
# =============================================================================
BASE_DIR = Path(__file__).resolve().parent          # .../DATA
SKU_DIR = BASE_DIR / "SKU"
SALES_DIR = BASE_DIR / "SALES"

RESULT_DIR = SALES_DIR / "RESULT"
LOGS_DIR = SALES_DIR / "LOGS"
MONTHLY_SALES_CORRECTED_DIR = SALES_DIR / "MONTHLY_SALES_CORRECTED"

# Дефолтні файли/папки
DEFAULT_SKU_MASTER_PATH = str(SKU_DIR / "sku_master.xlsx")
DEFAULT_MONTHLY_SALES_CORRECTED_DIR = str(MONTHLY_SALES_CORRECTED_DIR)
DEFAULT_CORRECTIONS_LOG_PATH = str(RESULT_DIR / "monthly_sales_corrections.log")

# Створюємо каталоги, якщо їх ще немає (щоб у Colab не падати на save/log)
RESULT_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)

SKU_SHEET_NAME = "SKU"
SKU_HEADER_NAME = "SKU (ключ, унікальний)"
NORM_UA_HEADER = "Normalized-UA"
MAPPING_SHEET_NAME = "Mapping"
MAPPING_WRONG_HEADER = "Неправильный"
MAPPING_RIGHT_HEADER = "Артикул"

SKU_ATTR_HEADERS = ["Name", "Brand", "Group1", "Group2"]

# Підсвітки
FILL_RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
FILL_GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
FILL_CLEAR = PatternFill()

# Month_sales (editable)
FILL_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
# Month_sales changed by corrections (differs from Mediana)
FILL_MONTH_CHANGED = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

# Нормалізація: прибрати пробіли (в т.ч. NBSP) + заміна схожої кирилиці
_WS_RE = re.compile(r"[\s\u00A0\u2007\u202F]+")
_C2L = {
    "А": "A", "В": "B", "Е": "E", "І": "I", "К": "K", "М": "M",
    "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
    "а": "A", "в": "B", "е": "E", "і": "I", "к": "K", "м": "M",
    "н": "H", "о": "O", "р": "P", "с": "C", "т": "T", "х": "X",
    "Ё": "E", "ё": "E",
    "Й": "I", "й": "I",
    "З": "3", "з": "3",
    "У": "Y", "у": "Y",
}


# =============================================================================
# Допоміжні функції
# =============================================================================
def _print_banner(title: str) -> None:
    line = "=" * 72
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(line)
    print(f"{title} | {ts}")
    print(line)


def normalize_symbol(value: object) -> str:
    """Нормалізує артикул:
    - прибирає ВСІ пробіли (включно з NBSP)
    - замінює візуально схожі кириличні символи на латиницю
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = _WS_RE.sub("", s)
    s = "".join(_C2L.get(ch, ch) for ch in s)
    return s


def _fmt_status(tag: str, use_color: bool = True) -> str:
    """Повертає короткий статус (OK/WRN/ERR) з опціональним ANSI-кольором."""
    if not use_color:
        return tag
    colors = {
        "OK": "\x1b[32m",   # green
        "WRN": "\x1b[33m",  # yellow
        "ERR": "\x1b[31m",  # red
    }
    reset = "\x1b[0m"
    return f"{colors.get(tag, '')}{tag}{reset}"



def _find_header_col(ws: openpyxl.worksheet.worksheet.Worksheet, header_name: str) -> Optional[int]:
    """Шукає колонку за точним заголовком у 1-му рядку."""
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is not None and str(v).strip() == header_name:
            return col
    return None


def _autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 5000) -> None:
    """М'яка авто-ширина колонок (обмежено по рядках, щоб не було повільно)."""
    try:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, min(ws.max_row, max_rows) + 1):
                v = ws.cell(r, col_idx).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)
    except Exception:
        pass



def _find_corrections_file(corrections_dir: str) -> Optional[str]:
    """Повертає шлях до першого .xlsx файлу в папці корекцій (без тимчасових ~$/)."""
    if not corrections_dir or not os.path.isdir(corrections_dir):
        return None
    xlsx = [f for f in os.listdir(corrections_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx:
        return None
    xlsx.sort()
    return os.path.join(corrections_dir, xlsx[0])


def _read_corrections_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Optional[float]]:
    """Читає лист корекцій: Symbol_normalized -> Month_sales (float/None)."""
    sym_col = _find_header_col(ws, NORMALIZED_COL_HEADER)
    ms_col = _find_header_col(ws, "Month_sales")
    if sym_col is None or ms_col is None:
        raise ValueError(
            f"У файлі корекцій на листі '{ws.title}' потрібні колонки "
            f"'{NORMALIZED_COL_HEADER}' та 'Month_sales'"
        )

    mp: Dict[str, Optional[float]] = {}
    for r in range(2, ws.max_row + 1):
        sym = normalize_symbol(ws.cell(r, sym_col).value)
        if not sym:
            continue
        mp[sym] = _to_number_or_none(ws.cell(r, ms_col).value)
    return mp


def _apply_month_sales_corrections(
    wb: openpyxl.Workbook,
    corrections_path: str,
    *,
    use_color: bool = True,
    log_path: str = DEFAULT_CORRECTIONS_LOG_PATH,
    norm_ua: Optional[Dict[str, float]] = None,
) -> None:
    """Застосовує корекції Month_sales з файлу до листів *-normalized у wb.

    - замінює Month_sales якщо Symbol_normalized знайдений у корекціях
    - якщо нове значення != Mediana -> заливка D червона, інакше жовта
    - якщо Symbol у корекціях відсутній на відповідному листі results -> пише в детальний лог
    """
    corr_wb = openpyxl.load_workbook(corrections_path, data_only=True)

    regions = ["UA", "KZ", "UZ"]
    summary = {r: {"replaced": 0, "changed": 0, "not_found": 0, "total_corr": 0} for r in regions}
    detail_lines: List[str] = []
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    detail_lines.append(f"[{ts}] Corrections file: {corrections_path}")

    for reg in regions:
        sheet = f"{reg}-normalized"
        if sheet not in corr_wb.sheetnames:
            continue
        corr_map = _read_corrections_sheet(corr_wb[sheet])
        summary[reg]["total_corr"] = len(corr_map)

        if sheet not in wb.sheetnames:
            summary[reg]["not_found"] += len(corr_map)
            for sym in sorted(corr_map.keys()):
                detail_lines.append(f"{reg}\tMISSING_RESULTS_SHEET\t{sym}")
            continue

        ws = wb[sheet]
        sym_col = _find_header_col(ws, NORMALIZED_COL_HEADER)
        med_col = _find_header_col(ws, "Mediana")
        ms_col = _find_header_col(ws, "Month_sales")
        if sym_col is None or med_col is None or ms_col is None:
            continue

        idx: Dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            sym = normalize_symbol(ws.cell(r, sym_col).value)
            if sym:
                idx[sym] = r

        for sym, new_ms in corr_map.items():
            rr = idx.get(sym)
            if rr is None:
                summary[reg]["not_found"] += 1
                detail_lines.append(f"{reg}\tNOT_FOUND_IN_RESULTS\t{sym}\tnew={new_ms}")
                continue

            old_ms = _to_number_or_none(ws.cell(rr, ms_col).value)
            applied_ms = new_ms
            if reg == "UA" and applied_ms is not None:
                k = (norm_ua or {}).get(sym)
                factor = float(k) if k not in (None, 0) else 1.0
                applied_ms = float(applied_ms) * factor
            ws.cell(rr, ms_col).value = applied_ms
            summary[reg]["replaced"] += 1

            med = _to_number_or_none(ws.cell(rr, med_col).value)
            cell = ws.cell(rr, ms_col)

            if (applied_ms is not None) and (med is not None) and (abs(applied_ms - med) > 1e-9):
                cell.fill = FILL_MONTH_CHANGED
                summary[reg]["changed"] += 1
            else:
                cell.fill = FILL_YELLOW

            if old_ms is not None and applied_ms is not None and abs(old_ms - applied_ms) > 1e-9:
                detail_lines.append(f"{reg}\tUPDATED\t{sym}\told={old_ms}\tnew={new_ms}\tmed={med}")

    try:
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        with open(log_path, "a", encoding="utf-8") as f:
            for ln in detail_lines:
                f.write(ln + "\n")
    except Exception:
        pass

    print(f"{_fmt_status('OK', use_color)}   | step=3 | corrections_file={os.path.basename(corrections_path)}")
    for reg in regions:
        s = summary[reg]
        if s["total_corr"] == 0 and s["replaced"] == 0 and s["not_found"] == 0:
            continue

        print(
            f"{_fmt_status('OK', use_color)}   | region={reg:<2} | corrections={s['total_corr']:<5} | applied={s['replaced']:<5} | changed={s['changed']:<5} | not_found={s['not_found']:<5}"
        )
    total_nf = sum(summary[r]["not_found"] for r in regions)
    if total_nf > 0:
        print(f"{_fmt_status('WRN', use_color)}  | step=3 | corrections_not_found_total={total_nf} | log={os.path.abspath(log_path)}")

def load_sku_master(
    sku_master_path: str,
) -> Tuple[Set[str], Dict[str, str], Dict[str, Dict[str, str]], Dict[str, float]]:
    """Завантажує:
    - sku_set (лист 'SKU', колонка SKU_HEADER_NAME)
    - mapping wrong->right (лист 'Mapping')
    - attrs: dict[SKU] -> {"Name":..,"Brand":..,"Group1":..,"Group2":..} (лист 'SKU')
    - norm_ua: dict[SKU] -> Normalized-UA (float) (лист 'SKU', колонка NORM_UA_HEADER)

    Всі SKU ключі приводяться до normalize_symbol().
    """
    def _to_float_local(x: object) -> Optional[float]:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        try:
            s = str(x).strip().replace(",", ".")
            if s == "":
                return None
            return float(s)
        except Exception:
            return None

    wb = openpyxl.load_workbook(sku_master_path, data_only=True)

    if SKU_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"У sku_master.xlsx немає листа '{SKU_SHEET_NAME}'")

    ws_sku = wb[SKU_SHEET_NAME]
    sku_col = _find_header_col(ws_sku, SKU_HEADER_NAME)
    if sku_col is None:
        raise ValueError(
            f"На листі '{SKU_SHEET_NAME}' не знайдено заголовок '{SKU_HEADER_NAME}' у 1-му рядку"
        )

    attr_cols: Dict[str, int] = {}
    for h in SKU_ATTR_HEADERS:
        c = _find_header_col(ws_sku, h)
        if c is not None:
            attr_cols[h] = c

    norm_ua_col = _find_header_col(ws_sku, NORM_UA_HEADER)

    sku_set: Set[str] = set()
    attrs: Dict[str, Dict[str, str]] = {}
    norm_ua: Dict[str, float] = {}

    for r in range(2, ws_sku.max_row + 1):
        sku = normalize_symbol(ws_sku.cell(r, sku_col).value)
        if not sku:
            continue
        sku_set.add(sku)

        row_attrs: Dict[str, str] = {}
        for h, cidx in attr_cols.items():
            v = ws_sku.cell(r, cidx).value
            row_attrs[h] = "" if v is None else str(v).strip()
        attrs[sku] = row_attrs

        if norm_ua_col is not None:
            k = _to_float_local(ws_sku.cell(r, norm_ua_col).value)
            if k is not None:
                norm_ua[sku] = float(k)

    mapping: Dict[str, str] = {}
    if MAPPING_SHEET_NAME in wb.sheetnames:
        ws_map = wb[MAPPING_SHEET_NAME]
        wrong_col = _find_header_col(ws_map, MAPPING_WRONG_HEADER)
        right_col = _find_header_col(ws_map, MAPPING_RIGHT_HEADER)
        if wrong_col is None or right_col is None:
            raise ValueError(
                f"На листі '{MAPPING_SHEET_NAME}' очікуються заголовки "
                f"'{MAPPING_WRONG_HEADER}' та '{MAPPING_RIGHT_HEADER}' у 1-му рядку"
            )

        for r in range(2, ws_map.max_row + 1):
            wrong = normalize_symbol(ws_map.cell(r, wrong_col).value)
            right = normalize_symbol(ws_map.cell(r, right_col).value)
            if wrong and right:
                mapping[wrong] = right

    return sku_set, mapping, attrs, norm_ua
def _to_number_or_none(x: object) -> Optional[float]:
    """Пробує конвертувати в float; якщо не виходить — None."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        s = str(x).strip().replace(",", ".")
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def _sum_cells(a: object, b: object) -> object:
    """Сума для числових клітинок, інакше — "краще" з двох значень."""
    an = _to_number_or_none(a)
    bn = _to_number_or_none(b)
    if an is None and bn is None:
        return a if a not in (None, "") else b
    if an is None:
        return bn
    if bn is None:
        return an
    return an + bn


# =============================================================================
# Arrivals integration (UA+/KZ+/UZ+)
# =============================================================================
def _load_arrivals_qty_by_region(wb: openpyxl.Workbook, reg: str) -> Dict[str, float]:
    """Читає лист '<REG>+' (якщо є) і повертає мапу:
    Symbol_normalized -> Quantity (float).

    Очікувані заголовки у 1-му рядку: Symbol, Quantity (Symbol_raw ігноруємо).
    """
    sheet = f"{reg}+"
    if sheet not in wb.sheetnames:
        return {}

    ws = wb[sheet]
    sym_col = _find_header_col(ws, "Symbol")
    qty_col = _find_header_col(ws, "Quantity")
    if sym_col is None or qty_col is None:
        # Лист є, але формат несподіваний — пропускаємо без падіння.
        return {}

    mp: Dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        sym = normalize_symbol(ws.cell(r, sym_col).value)
        if not sym:
            continue
        q = _to_number_or_none(ws.cell(r, qty_col).value)
        if q is None:
            continue
        mp[sym] = float(mp.get(sym, 0.0) + q)
    return mp

# =============================================================================
# Головна утиліта
# =============================================================================
def normalize_results_workbook(
    results_xlsx_path: str,
    sku_master_path: Optional[str] = None,
    corrections_dir: Optional[str] = None,
    *,
    use_color: bool = True,
    corrections_log_path: str = DEFAULT_CORRECTIONS_LOG_PATH,
    include_arrivals_added_col: bool = False,
) -> str:
    """Створює/оновлює *-normalized листи у results.xlsx і зберігає книгу."""
    _print_banner("STEP 3 | Normalize results workbook")

    sku_master_path = sku_master_path or DEFAULT_SKU_MASTER_PATH

    # --- Load SKU master (optional) ---
    sku_set: Set[str] = set()
    mapping: Dict[str, str] = {}
    sku_attrs: Dict[str, Dict[str, str]] = {}
    norm_ua: Dict[str, float] = {}
    try:
        if not os.path.exists(sku_master_path):
            raise FileNotFoundError(f"sku_master.xlsx не знайдено: {sku_master_path}")
        sku_set, mapping, sku_attrs, norm_ua = load_sku_master(sku_master_path)
    except Exception as exc:
        print(f"WARNING: SKU master не завантажився, підсвіток/валідації не буде: {exc}")

    if not os.path.exists(results_xlsx_path):
        raise FileNotFoundError(f"results.xlsx не знайдено: {results_xlsx_path}")

    wb = openpyxl.load_workbook(results_xlsx_path)

    # --- Завантажити (UA+/KZ+/UZ+) ,якщо вони там є (зроблено на 2-му кроці) ---
    arrivals_qty: Dict[str, Dict[str, float]] = {
        "UA": _load_arrivals_qty_by_region(wb, "UA"),
        "KZ": _load_arrivals_qty_by_region(wb, "KZ"),
        "UZ": _load_arrivals_qty_by_region(wb, "UZ"),
    }

    regions = [
        ("UA", "UA-sales", "UA-normalized"),
        ("KZ", "KZ-sales", "KZ-normalized"),
        ("UZ", "UZ-sales", "UZ-normalized"),
    ]

    for reg, sales_sheet, norm_sheet in regions:
        if sales_sheet not in wb.sheetnames:
            print(f"{_fmt_status('WRN', use_color)}  | region={reg:<2} | sheet={sales_sheet:<9} | reason=missing_sheet")
            continue

        ws_src = wb[sales_sheet]

        # replace normalized sheet
        if norm_sheet in wb.sheetnames:
            wb.remove(wb[norm_sheet])
        ws_dst = wb.create_sheet(norm_sheet)

        # headers
        headers = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
        header_keys = ["" if h is None else str(h).strip() for h in headers]

        # find Symbol header position (fallback A)
        symbol_key = SYMBOL_HEADER_NAME
        symbol_col_idx = 1
        for i, hk in enumerate(header_keys, start=1):
            if hk == symbol_key:
                symbol_col_idx = i
                break

        # ------------------------------------------------------------
        # Місячні колонки: зчитуємо з будь-яких назв і приводимо до янв..дек
        # ------------------------------------------------------------
        # мапа: канонічний місяць -> реальна назва колонки у джерелі
        month_col_by_canon: Dict[str, str] = {}
        for hk in header_keys:
            if hk in ("", symbol_key, "Stock", "TOTAL"):
                continue
            cm = _canon_month(hk)
            if cm and cm not in month_col_by_canon:
                month_col_by_canon[cm] = hk

        # фіксований порядок місяців для медіани/виводу
        month_keys: List[str] = REQUIRED_COLUMNS[1:]

        # --- build normalized rows with mapping/color ---
        rows_norm: List[Dict[str, object]] = []
        for r in range(2, ws_src.max_row + 1):
            row_vals = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
            row_vals = row_vals + [None] * max(0, len(headers) - len(row_vals))

            sym_val = row_vals[symbol_col_idx - 1] if symbol_col_idx - 1 < len(row_vals) else None

            sym_norm = normalize_symbol(sym_val)
            final_norm = sym_norm
            status_fill = None  # green/red/None

            if final_norm and sku_set:
                if final_norm in sku_set:
                    status_fill = None
                else:
                    mapped = mapping.get(final_norm)
                    if mapped and mapped in sku_set:
                        final_norm = mapped
                        status_fill = FILL_GREEN
                    else:
                        status_fill = FILL_RED

            values = {header_keys[i]: row_vals[i] for i in range(len(header_keys))}

            # Додаємо канонічні місяці (янв..дек) у словник значень
            for cm in month_keys:
                src_col = month_col_by_canon.get(cm)
                values[cm] = values.get(src_col) if src_col else None

            rows_norm.append({"symbol": sym_val, "sym_norm": final_norm, "fill": status_fill, "values": values})

        # --- aggregate by sym_norm for uniqueness ---
        agg: Dict[str, Dict[str, object]] = {}
        for item in rows_norm:
            key = item["sym_norm"] or ""
            if key == "":
                continue

            if key not in agg:
                agg[key] = {
                    "symbol": item["symbol"],
                    "sym_norm": key,
                    "fill": item["fill"],
                    "values": dict(item["values"]),
                }
            else:
                # merge fill: red beats green beats None
                prev_fill = agg[key]["fill"]
                cur_fill = item["fill"]
                if prev_fill == FILL_RED or cur_fill == FILL_RED:
                    agg[key]["fill"] = FILL_RED
                elif prev_fill == FILL_GREEN or cur_fill == FILL_GREEN:
                    agg[key]["fill"] = FILL_GREEN
                else:
                    agg[key]["fill"] = None

                prev_vals = agg[key]["values"]
                cur_vals = item["values"]
                for col in header_keys:
                    if col in ("", symbol_key):
                        continue
                    prev_vals[col] = _sum_cells(prev_vals.get(col), cur_vals.get(col))

        keys_sorted = sorted(agg.keys())

        def _get_attr(sku: str, key: str) -> str:
            return sku_attrs.get(sku, {}).get(key, "")

        red_keys = [k for k in keys_sorted if agg[k].get('fill') == FILL_RED]
        red_keys.sort(key=lambda x: x)

        ok_keys = [k for k in keys_sorted if agg[k].get('fill') != FILL_RED]
        ok_keys.sort(key=lambda x: (
            _get_attr(x, 'Name'),
            _get_attr(x, 'Brand'),
            _get_attr(x, 'Group1'),
            _get_attr(x, 'Group2'),
            x,
        ))

        ordered_keys = red_keys + ok_keys

        merged = max(0, len(rows_norm) - len(agg))
        green = sum(1 for rec in agg.values() if rec.get('fill') == FILL_GREEN)
        red = sum(1 for rec in agg.values() if rec.get('fill') == FILL_RED)

        # --- output headers order ---
        # Symbol, Symbol_normalized, Mediana, Month_sales, Stock, then the rest (excluding Symbol/Stock)
        # base_other: лишаємо інші колонки, але ВИКИДАЄМО сирі місячні заголовки,
        # щоб у виході не було дублю (довгі назви + янв..дек)
        raw_month_headers = set(month_col_by_canon.values())
        base_other = [
            h for h in header_keys
            if h not in ("", symbol_key, "Stock") and h not in raw_month_headers
        ]

        # Додаємо канонічні місяці у фіксованому порядку на початок вихідних колонок
        base_other = month_keys + base_other
        # TOTAL: має бути сумою продажів за 12 місяців (у *-normalized рахуємо з колонок янв..дек)
        # Ставимо TOTAL одразу після 12 місяців, навіть якщо в джерелі колонки TOTAL не було або вона була не в тому місці.
        if "TOTAL" in base_other:
            base_other = [h for h in base_other if h != "TOTAL"]
        base_other.insert(len(month_keys), "TOTAL")
        out_headers = [symbol_key, NORMALIZED_COL_HEADER, "Mediana", "Month_sales", "Stock"]
        if include_arrivals_added_col:
            out_headers.append("Arrivals_added")
        out_headers += base_other

        # Вставляємо атрибути SKU після TOTAL
        if "TOTAL" in out_headers:
            ti = out_headers.index("TOTAL") + 1
            for h in ["Name", "Brand", "Group1", "Group2"]:
                if h not in out_headers:
                    out_headers.insert(ti, h)
                    ti += 1

        for c, h in enumerate(out_headers, start=1):
            ws_dst.cell(1, c, h)

        # column indices (1-based)
        sym_norm_col = 2
        month_sales_col = 4

        out_r = 2
        for k in ordered_keys:
            rec = agg[k]
            vals = rec["values"]

            # UA-only: scale values by Normalized-UA from sku_master (column "Normalized-UA")
            # UA-only: scale values by Normalized-UA from sku_master (column "Normalized-UA")
            factor = 1.0
            if reg == "UA":
                k_ua = norm_ua.get(rec["sym_norm"])
                if k_ua not in (None, 0):
                    factor = float(k_ua)

            # median of month columns (already scaled for UA)
            nums: List[float] = []
            for mk in month_keys:
                n = _to_number_or_none(vals.get(mk))
                if n is not None:
                    nums.append(float(n) * factor)
            med = float(statistics.median(nums)) if nums else None

            # --- ARRIVALS: UA-normalized має брати Quantity * Normalized-UA ---
            arr_raw = arrivals_qty.get(reg, {}).get(rec["sym_norm"])  # як у листі UA+/KZ+/UZ+
            arr_out = (float(arr_raw) * factor) if (reg == "UA" and arr_raw is not None) else arr_raw

            # Stock: додаємо arrivals (сирі), а потім масштабуємо фактором -> еквівалентно множенню arrivals теж
            raw_stock = _sum_cells(vals.get("Stock"), arr_raw)
            stock_n = _to_number_or_none(raw_stock)
            stock_out = (float(stock_n) * factor) if (stock_n is not None) else raw_stock

            row_map: Dict[str, object] = {
                symbol_key: rec["symbol"],
                NORMALIZED_COL_HEADER: rec["sym_norm"],
                "Mediana": med,
                "Month_sales": med,
                "Stock": stock_out,
                **({"Arrivals_added": arr_out} if include_arrivals_added_col else {}),
            }

            # Атрибути SKU (тільки якщо SKU існує у sku_set)
            if rec["sym_norm"] in sku_set:
                row_map["Name"] = sku_attrs.get(rec["sym_norm"], {}).get("Name", "")
                row_map["Brand"] = sku_attrs.get(rec["sym_norm"], {}).get("Brand", "")
                row_map["Group1"] = sku_attrs.get(rec["sym_norm"], {}).get("Group1", "")
                row_map["Group2"] = sku_attrs.get(rec["sym_norm"], {}).get("Group2", "")

            for h in base_other:
                if reg == "UA" and h in month_keys:
                    n = _to_number_or_none(vals.get(h))
                    row_map[h] = (float(n) * factor) if n is not None else vals.get(h)
                else:
                    row_map[h] = vals.get(h)


            # TOTAL = сума продажів за 12 місяців (враховуючи UA множник, бо місяці вже записані з factor)
            total_nums: List[float] = []
            for mk in month_keys:
                tn = _to_number_or_none(row_map.get(mk))
                if tn is not None:
                    total_nums.append(float(tn))
            row_map["TOTAL"] = sum(total_nums) if total_nums else None

            for c, h in enumerate(out_headers, start=1):
                cell = ws_dst.cell(out_r, c, row_map.get(h))

                if c == sym_norm_col:
                    fill = rec.get("fill")
                    if fill is not None:
                        cell.fill = fill  # green/red
                    else:
                        cell.fill = FILL_CLEAR
                elif c == month_sales_col:
                    cell.fill = FILL_YELLOW

            out_r += 1

        _autosize_columns(ws_dst)

        in_rows = max(0, ws_src.max_row - 1)
        out_rows = len(agg)

        # arrivals -> Stock addition stats
        arr_map = arrivals_qty.get(reg, {})
        arrivals_applied = sum(
            1 for sku in agg.keys()
            if (sku in arr_map and (arr_map.get(sku) or 0.0) != 0.0)
        )

        print(
            f"{_fmt_status('OK', use_color)}   | region={reg:<2} | in={in_rows:<5} | unique={out_rows:<5} | merged={merged:<5} | green={green:<4} | red={red:<4} | arrivals_added={arrivals_applied:<4} | saved={norm_sheet}"
        )


    # --- apply Month_sales corrections (optional) ---
    corr_dir = corrections_dir or DEFAULT_MONTHLY_SALES_CORRECTED_DIR
    corr_path = _find_corrections_file(corr_dir)
    if corr_path:
        try:
            _apply_month_sales_corrections(
                wb,
                corr_path,
                use_color=use_color,
                log_path=corrections_log_path,
                norm_ua=norm_ua,
            )
        except Exception as exc:
            print(f"{_fmt_status('WRN', use_color)}  | step=3 | corrections_failed | {exc}")

    # --- reorder sheets at end ---
    desired_tail = ["UA+", "KZ+", "UZ+", "UA-normalized", "KZ-normalized", "UZ-normalized", "UA-sales", "KZ-sales", "UZ-sales"]
    others = [ws for ws in wb._sheets if ws.title not in desired_tail]
    tail = [wb[s] for s in desired_tail if s in wb.sheetnames]
    wb._sheets = others + tail

    wb.save(results_xlsx_path)
    print(f"{_fmt_status('OK', use_color)}   | step=3 | saved={os.path.abspath(results_xlsx_path)}")
    return os.path.abspath(results_xlsx_path)


def main() -> None:
    """CLI entrypoint.

    Типовий сценарій у Colab:
        !python DATA/normalize_results_workbook.py

    За потреби можна передати інші шляхи через аргументи.
    """
    import argparse

    parser = argparse.ArgumentParser(description="Крок 4: нормалізація results.xlsx та уніфікація місяців (янв..дек).")
    parser.add_argument(
        "--results",
        default=str(RESULT_DIR / "results.xlsx"),
        help="Шлях до results.xlsx (за замовчуванням: DATA/SALES/RESULT/results.xlsx).",
    )
    parser.add_argument(
        "--sku-master",
        default=DEFAULT_SKU_MASTER_PATH,
        help="Шлях до sku_master.xlsx (за замовчуванням: DATA/SKU/sku_master.xlsx).",
    )
    parser.add_argument(
        "--corrections-dir",
        default=DEFAULT_MONTHLY_SALES_CORRECTED_DIR,
        help="Папка з файлами корекцій місячних продажів (MONTHLY_SALES_CORRECTED).",
    )
    parser.add_argument(
        "--corrections-log",
        default=DEFAULT_CORRECTIONS_LOG_PATH,
        help="Файл логу для SKU, які не знайдено під час застосування корекцій.",
    )
    parser.add_argument(
        "--no-color",
        action="store_true",
        help="Вимкнути кольорові статуси в консолі.",
    )
    parser.add_argument(
        "--include-arrivals-added-col",
        action="store_true",
        help="Додати колонку Arrivals_added у *-normalized (якщо доступні аркуші UA+/KZ+/UZ+).",
    )

    args = parser.parse_args()

    normalize_results_workbook(
        results_xlsx_path=args.results,
        sku_master_path=args.sku_master,
        corrections_dir=args.corrections_dir,
        corrections_log_path=args.corrections_log,
        include_arrivals_added_col=args.include_arrivals_added_col,
        use_color=not args.no_color,
    )


if __name__ == "__main__":
    main()