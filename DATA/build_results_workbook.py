"""
build_results_workbook.py

Пакетний парсер Excel-звітів ПРОДАЖІВ (UA/KZ/UZ) з "плаваючими" заголовками.

2) Якщо евристики не змогли розпізнати шапку/місяці — підключаємо LLM (Gemini API) як fallback.

Важливі правила для обробки:
- Рядки без Symbol (артикула) НЕ переносимо у вихідний файл.
- Якщо у заголовках є "ОС" (після нормалізації -> "ос") — це Stock і регіон UZ визначається однозначно.
- Перед аналізом "розкриваємо" всі приховані колонки (hidden=False) і вже потім детектимо/рахуємо.
- TOTAL у вихідному файлі завжди перераховується як сума місяців.
  Якщо у вхідному файлі був власний Total — підсвічуємо mismatch червоним.

Модулі, що потрібні:
- openpyxl
- requests
- python-dotenv (імпорт: from dotenv import load_dotenv)

"""

from __future__ import annotations

import json
import datetime
import logging
import os
import re
import shutil
from dataclasses import dataclass, replace
from typing import Dict, List, Optional, Sequence, Tuple

import requests
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook import Workbook

# =============================================================================
# КРАСИВИЙ ВИВІД У КОНСОЛІ (ANSI кольори + банер)
# =============================================================================
def _print_banner(title: str) -> None:
    line = "=" * 72
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(line)
    print(f"{title} | {ts}")
    print(line)


def _fmt_status(tag: str, use_color: bool = True) -> str:
    """Повертає короткий статус (OK/WRN/ERR/FAIL) з опціональним ANSI-кольором."""
    if not use_color:
        return tag
    colors = {
        "OK": "\x1b[32m",    # green
        "WRN": "\x1b[33m",   # yellow
        "ERR": "\x1b[31m",   # red
        "FAIL": "\x1b[31m",  # red
    }
    reset = "\x1b[0m"
    return f"{colors.get(tag, '')}{tag}{reset}"



# =============================================================================
# ПАПКИ ПРОЄКТУ (усі всередині SALES/)
# =============================================================================
SALES_ROOT_DIR = "SALES"
SALES_REPORTS_DIR = os.path.join(SALES_ROOT_DIR, "SALES_REPORTS")
RECOGNIZED_SALES_REPORTS_DIR = os.path.join(SALES_ROOT_DIR, "RECOGNIZED_SALES_REPORTS")
FAILED_SALES_REPORTS_DIR = os.path.join(SALES_ROOT_DIR, "FAILED_SALES_REPORTS")
LOGS_DIR = os.path.join(SALES_ROOT_DIR, "LOGS")

# =============================================================================
# RESULTS (агрегація по регіонах)
# =============================================================================
RESULT_DIR = os.path.join(SALES_ROOT_DIR, "RESULT")
RESULTS_XLSX = os.path.join(RESULT_DIR, "results.xlsx")

# =============================================================================
# ЛОГІНГ
# =============================================================================
LOG_FILE = os.path.join(LOGS_DIR, "sales_parser.log")

# =============================================================================
# LLM (Gemini) налаштування
# =============================================================================
GEMINI_MODEL = "gemini-2.5-flash-lite"
GEMINI_ENDPOINT = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    f"{GEMINI_MODEL}:generateContent"
)

# Щоб не відправляти занадто багато в LLM:
LLM_MAX_ROWS = 30
# NOTE: LLM preview must be wide enough to include all month columns.
# Some vendor reports place months far to the right (30-80+ cols).
LLM_MAX_COLS = 120

# =============================================================================
# СКАН ЗВЕРХУ ФАЙЛУ
# =============================================================================
PREFERRED_SHEET_NAME = "Лист1"
HEADER_SCAN_LIMIT = 150
# Header scanning width. Some reports have many auxiliary columns before months.
MAX_COLS_SCAN = 200  # widened to avoid cutting off month columns
# ДОПУСК ДЛЯ FLOAT
# =============================================================================
TOL = 1e-6

# =============================================================================
# КОРОТКИЙ ВИВІД В КОНСОЛІ (1 рядок на файл)
# =============================================================================
FILENAME_COL_WIDTH = 56

# =============================================================================
# ПІДСВІТКА MISMATCH
# =============================================================================
FILL_RED = PatternFill("solid", fgColor="FFC7CE")


# =============================================================================
# СИНОНІМИ КЛЮЧОВИХ КОЛОНОК (для м'якого contains-пошуку)
# =============================================================================
SYMBOL_NAMES_RAW = [
"symbol",
    "sku",
    "item",
    "item no",
    "item number",
    "артикул",
    "арт",
    "код товара",
    "код",
    "№ каталога",
    "кат номер",
    "кат. номер",
    "каталоговый номер",
    "каталожный номер",
    "номер каталога",
    "каталог",
    "товар",
    "product",
    "part number",
    "part no",
    ]

TOTAL_NAMES_RAW = [
    "total",
    "всього",
    "всего",
    "итого",
    "разом",
    "sum",
    "suma",
    "grand total",
    "підсумок",
    "подсумок",
]

STOCK_NAMES_RAW = [
    "stock",
    "залишок",
    "залишки",
    "остаток",
    "остатки",
    "на складі",
    "на складе",
    "склад",
    "warehouse",
    "ending balance",
    "balance",
    "saldo",
    # ✅ "ОС" у шапці = Stock
    "ос",
]

# =============================================================================
# НОРМАЛІЗАЦІЯ ТЕКСТУ
# =============================================================================
_PUNCT_RE = re.compile(r"[^0-9a-zа-яіїєёґ\s]+", flags=re.IGNORECASE)


def setup_logger() -> logging.Logger:
    """
    Налаштовує логер:
    - детальні логи пишемо у SALES/LOGS/sales_parser.log
    - у консоль не сиплемо (там буде 1 рядок через print у main()).
    """
    logger = logging.getLogger("sales_parser")
    logger.setLevel(logging.DEBUG)

    # У ноутбуці/перезапусках не дублюємо handlers
    if logger.handlers:
        return logger

    os.makedirs(LOGS_DIR, exist_ok=True)

    fmt = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger


LOGGER = setup_logger()


def ensure_dirs() -> None:
    """Створює потрібні папки проєкту."""
    os.makedirs(SALES_REPORTS_DIR, exist_ok=True)
    os.makedirs(RECOGNIZED_SALES_REPORTS_DIR, exist_ok=True)
    os.makedirs(FAILED_SALES_REPORTS_DIR, exist_ok=True)
    os.makedirs(LOGS_DIR, exist_ok=True)


def ensure_results_dir() -> None:
    """Створює папку SALES/RESULT, якщо її ще немає."""
    os.makedirs(RESULT_DIR, exist_ok=True)


def read_cleaned_rows(cleaned_xlsx_path: str) -> Tuple[List[object], List[List[object]]]:
    """Зчитує заголовок + всі рядки з нормалізованого файлу.

    Повертає:
      headers: список значень першого рядка
      rows:    список рядків (починаючи з 2-го)

    Примітка: це лише значення (data_only=True), без стилів.
    """
    wb = openpyxl.load_workbook(cleaned_xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    rows: List[List[object]] = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # пропускаємо пусті (на всяк випадок)
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
            continue
        rows.append(row_vals)

    return headers, rows


def align_rows_to_headers(
    target_headers: List[object],
    source_headers: List[object],
    source_rows: List[List[object]],
) -> List[List[object]]:
    """Вирівнює source_rows під target_headers за назвою колонки.

    Якщо заголовок не знайдений — значення буде None.
    Якщо в source є колонки, яких немає в target — вони ігноруються.
    """
    tgt = ["" if h is None else str(h).strip() for h in target_headers]
    src = ["" if h is None else str(h).strip() for h in source_headers]

    src_index = {name: i for i, name in enumerate(src)}

    aligned: List[List[object]] = []
    for row in source_rows:
        new_row: List[object] = [None] * len(tgt)
        for j, name in enumerate(tgt):
            i = src_index.get(name)
            if i is not None and i < len(row):
                new_row[j] = row[i]
        aligned.append(new_row)
    return aligned


def upsert_results_workbook(region_rows: Dict[str, Tuple[List[object], List[List[object]]]]) -> str:
    """Створює/оновлює SALES/RESULT/results.xlsx з 3 листами.

    Якщо книга вже існує — листи UA-sales/KZ-sales/UZ-sales замінюються.

    region_rows: {
      "UA": (headers, rows),
      "KZ": (headers, rows),
      "UZ": (headers, rows)
    }

    Повертає шлях до results.xlsx.
    """
    ensure_results_dir()

    if os.path.exists(RESULTS_XLSX):
        wb = openpyxl.load_workbook(RESULTS_XLSX)
    else:
        wb = Workbook()
        # прибираємо дефолтний лист
        if wb.sheetnames:
            ws0 = wb[wb.sheetnames[0]]
            wb.remove(ws0)

    for region, sheet_title in [("UA", "UA-sales"), ("KZ", "KZ-sales"), ("UZ", "UZ-sales")]:
        if sheet_title in wb.sheetnames:
            wb.remove(wb[sheet_title])
        ws = wb.create_sheet(sheet_title)

        headers, rows = region_rows.get(region, ([], []))
        if not headers:
            # Якщо для регіону нічого не розпізнали — лишаємо порожній лист з мінімальним повідомленням.
            ws.cell(1, 1, "No data")
            continue

        # Заголовок
        for c, h in enumerate(headers, start=1):
            ws.cell(1, c, h)

        # Дані
        out_r = 2
        for row_vals in rows:
            for c, v in enumerate(row_vals, start=1):
                ws.cell(out_r, c, v)
            out_r += 1

        # Невелике автопідлаштування ширини (грубо, але достатньо)
        try:
            for col_idx in range(1, len(headers) + 1):
                max_len = 0
                for r in range(1, min(ws.max_row, 2000) + 1):
                    v = ws.cell(r, col_idx).value
                    if v is None:
                        continue
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)
        except Exception:
            # не критично
            pass

    wb.save(RESULTS_XLSX)
    return RESULTS_XLSX



def pick_sheet_name(wb: openpyxl.Workbook) -> str:
    """
    Обирає лист для обробки:
    - якщо Arkusz1 існує — беремо його
    - інакше — перший лист
    """
    return PREFERRED_SHEET_NAME if PREFERRED_SHEET_NAME in wb.sheetnames else wb.sheetnames[0]


def unmerge_all(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Розмерджує всі merged-комірки та копіює значення верхньої-лівої
    комірки на весь блок (щоб заголовки було видно в кожній клітинці).
    """
    for merged_range in list(ws.merged_cells.ranges):
        min_r, min_c, max_r, max_c = merged_range.bounds
        top_left_value = ws.cell(min_r, min_c).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                ws.cell(r, c).value = top_left_value


def unhide_all_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_cols: int = 400) -> None:
    """
    Спочатку відображаємо всі приховані стовбчики.

    Ми не знаємо, скільки там колонок реально, тому ставимо обмеження max_cols.
    """
    max_col = min(ws.max_column, max_cols)
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].hidden = False


def normalize_text(value: object) -> str:
    """
    Нормалізує текст для порівняння заголовків:
    - None -> ""
    - lower
    - NBSP -> пробіл
    - прибирає пунктуацію
    - стискає пробіли
    """
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().lower()
    text = _PUNCT_RE.sub(" ", text)
    text = " ".join(text.split())
    return text


def build_keys(raw: Sequence[str]) -> List[str]:
    """
    Готує ключі для contains-match:
    - нормалізує
    - прибирає дублікати
    - сортує за довжиною (довші ключі першими)
    """
    uniq = {normalize_text(x) for x in raw if normalize_text(x)}
    return sorted(uniq, key=len, reverse=True)


SYMBOL_KEYS = build_keys(SYMBOL_NAMES_RAW)
TOTAL_KEYS = build_keys(TOTAL_NAMES_RAW)
STOCK_KEYS = build_keys(STOCK_NAMES_RAW)


def contains_any(text: str, keys: Sequence[str]) -> bool:
    """True, якщо text містить хоча б один ключ (substring)."""
    return bool(text) and any(k in text for k in keys)


def build_header_map(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    max_cols_scan: int = MAX_COLS_SCAN,
) -> Dict[int, str]:
    """Формує мапу заголовків {col_idx: normalized_header} у заданому рядку."""
    header_map: Dict[int, str] = {}
    max_col = min(ws.max_column, max_cols_scan)
    for c in range(1, max_col + 1):
        txt = normalize_text(ws.cell(header_row, c).value)
        if txt:
            header_map[c] = txt
    return header_map

def _is_month_header(norm_header: str) -> bool:
    """Heuristic: чи схожий заголовок на місяць продажів.

    ВАЖЛИВО: має працювати і для скорочень (Янв/Фев/Мар/Апр/Авг/Сен/Окт/Ноя/Дек),
    тому базуємось на month_index_from_header(), а не на окремому списку токенів.
    """
    if not norm_header:
        return False

    # Не плутаємо місяці зі складом/остатками
    if "остатк" in norm_header or "склад" in norm_header or "stock" in norm_header:
        return False

    return month_index_from_header(norm_header) is not None




def find_first_contains(header_map: Dict[int, str], keys: Sequence[str]) -> Optional[int]:
    """Повертає перший col_idx, де заголовок містить один з keys."""
    for col_idx, header in header_map.items():
        if contains_any(header, keys):
            return col_idx
    return None


def as_number(value: object) -> float:
    """Перетворює value на float, якщо це число. Інакше 0.0"""
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def safe_copy_to_failed(path: str) -> str:
    """Копіює файл у FAILED папку та повертає шлях призначення."""
    dst = os.path.join(FAILED_SALES_REPORTS_DIR, os.path.basename(path))
    shutil.copy2(path, dst)
    return dst


# =============================================================================
# ДЕТЕКЦІЯ МІСЯЦІВ У ЗАГОЛОВКАХ (RU/UA/EN + числові)
# =============================================================================
RU_MONTHS = {
    1: ["январ", "янв"],
    2: ["феврал", "февр", "фев"],
    3: ["март", "мар"],
    4: ["апрел", "апр"],
    5: ["май"],
    6: ["июн"],
    7: ["июл"],
    8: ["август", "авг"],
    9: ["сентябр", "сен", "сент"],
    10: ["октябр", "окт"],
    11: ["ноябр", "ноя"],
    12: ["декабр", "дек"],
}

UA_MONTHS = {
    1: ["січ", "сiч"],
    2: ["лют"],
    3: ["берез", "бер"],
    4: ["квіт", "квит"],
    5: ["трав"],
    6: ["черв"],
    7: ["лип"],
    8: ["серп"],
    9: ["верес", "вер"],
    10: ["жовт"],
    11: ["листоп", "лист"],
    12: ["груд"],
}

EN_MONTHS = {
    1: ["jan", "january"],
    2: ["feb", "february"],
    3: ["mar", "march"],
    4: ["apr", "april"],
    5: ["may"],
    6: ["jun", "june"],
    7: ["jul", "july"],
    8: ["aug", "august"],
    9: ["sep", "sept", "september"],
    10: ["oct", "october"],
    11: ["nov", "november"],
    12: ["dec", "december"],
}

MONTH_NUM_RE = re.compile(r"(?<!\d)(0?[1-9]|1[0-2])(?!\d)")


def month_index_from_header(header: str) -> Optional[int]:
    """
    Повертає номер місяця 1..12, якщо заголовок схожий на місяць.
    Інакше None.
    """
    if not header:
        return None
    h = header
    for idx, stems in RU_MONTHS.items():
        if any(stem in h for stem in stems):
            return idx
    for idx, stems in UA_MONTHS.items():
        if any(stem in h for stem in stems):
            return idx
    for idx, stems in EN_MONTHS.items():
        if any(stem in h for stem in stems):
            return idx

    m = MONTH_NUM_RE.search(h)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return num
    return None


def detect_region(filename: str, header_map: Dict[int, str], ws: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> str:
    """
    Визначає регіон (UA/KZ/UZ/UNK) для префікса вихідного файлу.

    Пріоритет:
    1) Якщо у заголовках є "ос" (ОС) як ОКРЕМЕ СЛОВО => Stock => регіон UZ однозначно.
       (За можливості шукаємо маркер не лише у header_row, а й у перших рядках аркуша,
       щоб не ламатись, коли LLM/евристики промахнулися з header_row.)
    2) Маркери в назві файлу.
    3) Визначення по мові місяців у заголовках.
    """
    os_re = re.compile(r"(?<!\w)ос(?!\w)", flags=re.IGNORECASE)

    def _has_os(values) -> bool:
        return any(os_re.search((v or "").strip()) for v in values)

    # 1) "ОС" => UZ (спершу у header_map)
    if _has_os(header_map.values()):
        return "UZ"

    # 1b) Якщо маємо worksheet — пошукаємо "ОС" у перших рядках, незалежно від header_row
    if ws is not None:
        max_row = min(getattr(ws, "max_row", 0) or 0, 20) or 20
        max_col = min(getattr(ws, "max_column", 0) or 0, MAX_COLS_SCAN) or MAX_COLS_SCAN
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                txt = normalize_text(ws.cell(r, c).value)
                if txt:
                    row_vals.append(txt)
            if row_vals and _has_os(row_vals):
                return "UZ"

    up = filename.upper()
    if "UA" in up or "ЮА" in up:
        return "UA"
    if "KZ" in up:
        return "KZ"
    if "UZ" in up:
        return "UZ"

    text_all = " ".join(header_map.values())
    if any(stem in text_all for stems in UA_MONTHS.values() for stem in stems):
        return "UA"
    if any(stem in text_all for stems in RU_MONTHS.values() for stem in stems):
        return "KZ"
    if any(stem in text_all for stems in EN_MONTHS.values() for stem in stems):
        return "UZ"

    return "UNK"


# =============================================================================
# ПОШУК HEADER ROW СКОРИНГОМ (евристики)
# =============================================================================
def score_row_as_header(header_map: Dict[int, str]) -> float:
    """
    Оцінює, наскільки рядок схожий на шапку.
    """
    if not header_map:
        return -1.0

    has_symbol = find_first_contains(header_map, SYMBOL_KEYS) is not None
    has_total = find_first_contains(header_map, TOTAL_KEYS) is not None

    # Stock: окремо врахуємо "ос"
    has_stock = find_first_contains(header_map, STOCK_KEYS) is not None or any(
        re.search(r"(?<!\w)ос(?!\w)", (h or "").strip(), flags=re.IGNORECASE) for h in header_map.values()
    )

    month_count = sum(1 for h in header_map.values() if month_index_from_header(h) is not None)

    score = 0.0
    score += 10.0 if has_symbol else 0.0
    score += 2.0 * float(month_count)
    score += 3.0 if has_total else 0.0
    score += 2.0 if has_stock else 0.0

    if len(header_map) < 3:
        score -= 5.0

    return score


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """
    Знаходить header row у перших HEADER_SCAN_LIMIT рядках за максимальним score.
    """
    max_row = min(ws.max_row, HEADER_SCAN_LIMIT)
    best_row = 0
    best_score = -1.0

    for r in range(1, max_row + 1):
        hm = build_header_map(ws, r)
        sc = score_row_as_header(hm)
        if sc > best_score:
            best_score = sc
            best_row = r

    if best_score < 12.0:
        raise ValueError("Не вдалося надійно знайти рядок заголовків (низький score).")

    LOGGER.debug("Header detection (heuristic): best_row=%s best_score=%s", best_row, best_score)
    return best_row


# =============================================================================
# LLM fallback: формуємо компактний "зріз" таблиці і просимо JSON з колонками
# =============================================================================
def _sheet_preview_as_tsv(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    max_rows: int = LLM_MAX_ROWS,
    max_cols: int = LLM_MAX_COLS,
) -> str:
    """
    Формує текстову прев'юшку листа у форматі TSV:
    - рядки 1..max_rows
    - колонки 1..max_cols
    """
    rows_out: List[str] = []
    max_r = min(ws.max_row, max_rows)
    max_c = min(ws.max_column, max_cols)

    for r in range(1, max_r + 1):
        row_vals: List[str] = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None:
                row_vals.append("")
            else:
                s = str(v).replace("\t", " ").replace("\n", " ").strip()
                row_vals.append(s)
        rows_out.append("\t".join(row_vals))
    return "\n".join(rows_out)


def _extract_json_object(text: str) -> Dict:
    """
    Витягує JSON-об'єкт з відповіді моделі.

    Моделі інколи додають пояснення навколо JSON, тому:
    - шукаємо перший '{' і останній '}' і пробуємо json.loads
    """
    if not text:
        raise ValueError("Порожня відповідь LLM.")
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("Не знайдено JSON у відповіді LLM.")
    candidate = text[start : end + 1]
    return json.loads(candidate)


def gemini_detect_structure(
    ws_struct: openpyxl.worksheet.worksheet.Worksheet,
    filename: str,
) -> Dict:
    """
    Викликає Gemini API для "розумного" визначення структури.

    Повертає dict з ключами:
      header_row: int
      symbol_col: int
      month_cols: list[int]
      total_col: int|null
      stock_col: int|null
      region_hint: str|null
    """
    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("Не знайдено GEMINI_API_KEY у .env або змінних середовища.")

    preview = _sheet_preview_as_tsv(ws_struct)

    prompt = f"""
Ти — помічник з аналізу Excel-таблиць продажів.

Ось прев'юшка листа (TSV, табами розділено колонки). У файлі можуть бути зміщені заголовки.
Твоє завдання: визначити структуру ТАБЛИЦІ ПРОДАЖІВ.

Потрібно повернути ЛИШЕ валідний JSON такого вигляду:
{{
  "header_row": <int>,                 // 1-based номер рядка з заголовками
  "symbol_col": <int>,                 // 1-based номер колонки з артикулом/Symbol
  "month_cols": [<int>, ...],          // 1-based колонки місяців у правильному порядку (1..12)
  "total_col": <int або null>,         // якщо є колонка Total у файлі
  "stock_col": <int або null>,         // якщо є колонка Stock/ОС/залишок
  "region_hint": "<UA|KZ|UZ|UNK>"       // якщо можеш, підкажи регіон
}}

Правила:
- month_cols має містити тільки колонки місяців (до 12-ти).
- Якщо Total відсутній, total_col має бути null.
- Якщо Stock відсутній, stock_col має бути null.
- Якщо в заголовках є "ОС" — це Stock і region_hint = "UZ".
- Не пиши жодних пояснень, лише JSON.

Filename: {filename}

TSV:
{preview}
""".strip()

    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0,
            "topP": 0.8,
            "topK": 20,
            "maxOutputTokens": 800,
            # ✅ важливо, щоб API повертав JSON, а не "ось ваш JSON..."
            "responseMimeType": "application/json",
        },
    }

    headers = {
        "Content-Type": "application/json",
        "x-goog-api-key": api_key,
    }

    MODELS_FALLBACK = [
        "gemini-2.5-flash-lite",  # основна для структури/JSON
        "gemini-2.5-flash",
        "gemini-2.5-pro",
    ]

    LOGGER.debug("LLM fallback: sending preview rows=%s cols=%s", LLM_MAX_ROWS, LLM_MAX_COLS)

    last_error = None
    data = None

    for model in MODELS_FALLBACK:
        endpoint = (
            "https://generativelanguage.googleapis.com/v1beta/models/"
            f"{model}:generateContent"
        )
        try:
            resp = requests.post(endpoint, headers=headers, json=payload, timeout=60)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")

            data = resp.json()

            # ✅ Перевірка, що є очікуваний шлях до тексту (або JSON)
            if "candidates" not in data or not data["candidates"]:
                raise RuntimeError(f"No candidates in response: keys={list(data.keys())}")

            # Якщо responseMimeType спрацює, модель може повернути текст як JSON-рядок або вже JSON.
            # Твій код нижче дістане "text" і _extract_json_object впорається.
            break

        except Exception as e:
            last_error = e
            LOGGER.warning("Gemini model failed: %s (%s)", model, e)
            continue

    if data is None:
        raise ValueError(f"All Gemini models failed. Last error: {last_error}")

    # Типова відповідь: candidates[0].content.parts[0].text
    try:
        text = data["candidates"][0]["content"]["parts"][0]["text"]
    except Exception as exc:
        raise ValueError(
            f"Незрозумілий формат відповіді Gemini: {exc}. Response keys: {list(data.keys())}"
        )

    result = _extract_json_object(text)
    return result


# =============================================================================
# ДЕТЕКЦІЯ СТРУКТУРИ SALES-ЗВІТУ
# =============================================================================
@dataclass(frozen=True)
class SalesDetection:
    sheet_name: str
    header_row: int
    col_symbol: int
    month_cols: List[int]
    col_total: Optional[int]
    col_stock: Optional[int]
    region: str
    used_llm: bool


def _detect_sales_structure_heuristic(
    ws_struct: openpyxl.worksheet.worksheet.Worksheet,
    filename: str,
) -> SalesDetection:
    """
    Евристичний детект структури.
    Якщо не вийшло — кидає ValueError (і тоді можна спробувати LLM).
    """
    header_row = find_header_row(ws_struct)
    # Use full width for LLM post-processing (months may be far to the right).
    header_map = build_header_map(ws_struct, header_row, max_cols_scan=ws_struct.max_column)

    col_symbol = find_first_contains(header_map, SYMBOL_KEYS)
    if col_symbol is None:
        raise ValueError("Не знайдено колонку Symbol/Артикул (обов’язкова).")

    col_total = find_first_contains(header_map, TOTAL_KEYS)

    # Stock + спеціальне правило для "ос"
    col_stock = find_first_contains(header_map, STOCK_KEYS)
    if col_stock is None:
        for col_idx, hdr in header_map.items():
            if os_re.search((hdr or "").strip()):
                col_stock = col_idx
                break

    # Місяці
    month_pairs: List[Tuple[int, int]] = []
    for col_idx, hdr in header_map.items():
        m_idx = month_index_from_header(hdr)
        if m_idx is not None:
            month_pairs.append((m_idx, col_idx))
    month_pairs.sort(key=lambda x: x[0])
    month_cols = [col for _, col in month_pairs]

    if not month_cols:
        raise ValueError("Не знайдено жодної колонки місяця (евристики).")

    region = detect_region(filename, header_map, ws_struct)

    return SalesDetection(
        sheet_name=ws_struct.title,
        header_row=header_row,
        col_symbol=col_symbol,
        month_cols=month_cols,
        col_total=col_total,
        col_stock=col_stock,
        region=region,
        used_llm=False,
    )


def _detect_sales_structure_llm(
    ws_struct: openpyxl.worksheet.worksheet.Worksheet,
    filename: str,
) -> SalesDetection:
    """
    LLM fallback.
    Тут ми беремо JSON-відповідь, валідуємо, і повертаємо SalesDetection.
    """
    llm = gemini_detect_structure(ws_struct, filename)

    header_row = int(llm["header_row"])
    col_symbol = int(llm["symbol_col"])
    month_cols = [int(x) for x in llm.get("month_cols", [])]

    total_col_raw = llm.get("total_col")
    stock_col_raw = llm.get("stock_col")

    col_total = int(total_col_raw) if isinstance(total_col_raw, int) else None
    col_stock = int(stock_col_raw) if isinstance(stock_col_raw, int) else None

    # Додаткова перевірка "ОС" (беремо всю ширину аркуша)
    header_map = build_header_map(ws_struct, header_row, max_cols_scan=ws_struct.max_column)
    os_re = re.compile(r"(?<!\w)ос(?!\w)", flags=re.IGNORECASE)
    if any(os_re.search((h or "").strip()) for h in header_map.values()):
        # якщо "ОС" присутній, але stock_col не дали — спробуємо знайти
        if col_stock is None:
            for col_idx, hdr in header_map.items():
                if os_re.search((hdr or "").strip()):
                    col_stock = col_idx
                    break

    region_hint = llm.get("region_hint")
    # Якщо LLM не впевнена (UNK/порожньо) — добиваємо евристиками по маркерам (наприклад "ОС" => UZ)
    if not region_hint or str(region_hint).strip().upper() == "UNK":
        region_hint = detect_region(filename, header_map, ws_struct)
    region = str(region_hint).upper()
    if region not in {"UA", "KZ", "UZ", "UNK"}:
        region = "UNK"

    # Мінімальна валідація
    if header_row < 1 or col_symbol < 1:
        raise ValueError("LLM повернула некоректні індекси (header_row/symbol_col).")
    if not month_cols:
        raise ValueError("LLM не повернула month_cols (не можу обробити sales-звіт).")

    return SalesDetection(
        sheet_name=ws_struct.title,
        header_row=header_row,
        col_symbol=col_symbol,
        month_cols=month_cols,
        col_total=col_total,
        col_stock=col_stock,
        region=region,
        used_llm=True,
    )


def detect_sales_structure(
    ws_struct: openpyxl.worksheet.worksheet.Worksheet,
    filename: str,
) -> SalesDetection:
    """
    Головна функція детекції структури:
    1) пробуємо евристику
    2) якщо fail — пробуємо LLM fallback
    """
    try:
        det = _detect_sales_structure_heuristic(ws_struct, filename)
        LOGGER.debug("Structure detected by heuristic.")
    except Exception as exc:
        LOGGER.warning("Heuristic detection failed: %s. Trying LLM fallback...", exc)
        det = _detect_sales_structure_llm(ws_struct, filename)
        LOGGER.debug("Structure detected by LLM fallback.")

    # --- SANITY FIXES ---
    if det.col_total is not None and det.col_total in det.month_cols:
        det = replace(det, month_cols=[c for c in det.month_cols if c != det.col_total])
        LOGGER.debug("Sanity: removed col_total=%s from month_cols", det.col_total)

    # Build header map across the full sheet width; otherwise month columns beyond
    # MAX_COLS_SCAN would be dropped by the sanity filter below.
    header_map = build_header_map(ws_struct, det.header_row, max_cols_scan=ws_struct.max_column)
    det = replace(det, month_cols=[c for c in det.month_cols if _is_month_header(normalize_text(header_map.get(c, "")))])

    return det


# =============================================================================
# ОБРОБКА ОДНОГО ФАЙЛУ
# =============================================================================
def build_output_filename(original_filename: str, region: str) -> str:
    """UA_<base>_sales_recognized.xlsx"""
    base = original_filename[:-5] if original_filename.lower().endswith(".xlsx") else original_filename
    return f"{region}_{base}_sales_recognized.xlsx"


def clean_one_sales_file(input_path: str, output_path_tmp: str) -> Tuple[int, int, str, bool]:
    """
    Обробляє один файл і зберігає у тимчасове ім’я (потім перейменуємо по region).

    Повертає:
        checked_rows: скільки рядків перенесли у вихід
        mismatches:   mismatch-рядки (лише якщо був оригінальний Total)
        region:       UA/KZ/UZ/UNK
        used_llm:     чи використовували LLM fallback
    """
    filename = os.path.basename(input_path)
    LOGGER.info("START file=%s", input_path)

    wb_struct = openpyxl.load_workbook(input_path, data_only=False)
    sheet_name = pick_sheet_name(wb_struct)
    ws_struct = wb_struct[sheet_name]

    # ✅ Порядок: розмерджити -> показати всі колонки -> детектити
    unmerge_all(ws_struct)
    unhide_all_columns(ws_struct)

    wb_val = openpyxl.load_workbook(input_path, data_only=True)
    ws_val = wb_val[sheet_name]

    det = detect_sales_structure(ws_struct, filename)
    # Full-width header map is required to correctly label month headers that are
    # located beyond MAX_COLS_SCAN.
    header_map = build_header_map(ws_struct, det.header_row, max_cols_scan=ws_struct.max_column)

    # Заголовки виходу: Symbol + місяці + TOTAL (Stock не виводимо)
    out_headers: List[str] = ["Symbol"]
    for col_idx in det.month_cols:
        out_headers.append(header_map.get(col_idx, f"month_{col_idx}"))
    out_headers.append("TOTAL")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = det.sheet_name

    for j, h in enumerate(out_headers, start=1):
        out_ws.cell(1, j, h)

    checked = 0
    mismatches = 0
    out_row = 2

    for r in range(det.header_row + 1, ws_struct.max_row + 1):
        if ws_struct.row_dimensions[r].hidden:
            continue

        # ✅ Рядки без Symbol видаляємо
        symbol_val = ws_val.cell(r, det.col_symbol).value
        symbol_txt = "" if symbol_val is None else str(symbol_val).strip()
        if not symbol_txt:
            continue

        month_values: List[float] = []
        computed_total = 0.0
        for c in det.month_cols:
            v = as_number(ws_val.cell(r, c).value)
            month_values.append(v)
            computed_total += v

        if det.col_total is not None:
            original_total = as_number(ws_val.cell(r, det.col_total).value)
            row_mismatch = abs(computed_total - original_total) > TOL
            if row_mismatch:
                mismatches += 1
                LOGGER.debug(
                    "Mismatch row=%s computed_total=%s original_total=%s diff=%s",
                    r, computed_total, original_total, computed_total - original_total
                )
        else:
            row_mismatch = False

        col_out = 1
        out_ws.cell(out_row, col_out, symbol_txt)
        col_out += 1

        for v in month_values:
            out_ws.cell(out_row, col_out, v)
            col_out += 1
        out_ws.cell(out_row, col_out, computed_total)

        if row_mismatch:
            for c in range(1, len(out_headers) + 1):
                out_ws.cell(out_row, c).fill = FILL_RED

        checked += 1
        out_row += 1

    out_wb.save(output_path_tmp)

    LOGGER.info(
        "DONE file=%s region=%s used_llm=%s checked=%s mismatches=%s saved=%s",
        input_path, det.region, det.used_llm, checked, mismatches, output_path_tmp
    )
    return checked, mismatches, det.region, det.used_llm


# =============================================================================
# BATCH-ОБРОБКА ПАПКИ
# =============================================================================
def build_results_workbook(
    sales_root_dir: str = SALES_ROOT_DIR,
    sales_reports_dir: str = SALES_REPORTS_DIR,
    recognized_dir: str = RECOGNIZED_SALES_REPORTS_DIR,
    failed_dir: str = FAILED_SALES_REPORTS_DIR,
    logs_dir: str = LOGS_DIR,
    results_dir: str = RESULT_DIR,
    results_xlsx: str = RESULTS_XLSX,
    *,
    use_color: bool = True,
) -> str:
    """
    Обробляє всі .xlsx у SALES/SALES_REPORTS.

    Консоль: 1 рядок на файл
    Деталі: SALES/LOGS/sales_parser.log
    """

    # Дозволяємо викликати утиліту з інших місць/ноутбуків з кастомними шляхами
    global SALES_ROOT_DIR, SALES_REPORTS_DIR, RECOGNIZED_SALES_REPORTS_DIR, FAILED_SALES_REPORTS_DIR, LOGS_DIR
    global RESULT_DIR, RESULTS_XLSX

    SALES_ROOT_DIR = sales_root_dir
    SALES_REPORTS_DIR = sales_reports_dir
    RECOGNIZED_SALES_REPORTS_DIR = recognized_dir
    FAILED_SALES_REPORTS_DIR = failed_dir
    LOGS_DIR = logs_dir
    RESULT_DIR = results_dir
    RESULTS_XLSX = results_xlsx

    _print_banner("STEP 1 | Build results workbook")

    ensure_dirs()
    LOGGER.info("======== NEW RUN ========")

    files = [
        f for f in os.listdir(SALES_REPORTS_DIR)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]

    total_files = len(files)
    success_files = 0
    llm_used_files = 0

    # Агрегація результатів для results.xlsx
    region_headers: Dict[str, List[object]] = {}
    region_rows: Dict[str, List[List[object]]] = {"UA": [], "KZ": [], "UZ": []}

    for filename in files:
        src = os.path.join(SALES_REPORTS_DIR, filename)

        tmp_name = build_output_filename(filename, "UNK")
        tmp_dst = os.path.join(RECOGNIZED_SALES_REPORTS_DIR, tmp_name)

        try:
            checked, mism, region, used_llm = clean_one_sales_file(src, tmp_dst)

            success_files += 1
            if used_llm:
                llm_used_files += 1

            final_name = build_output_filename(filename, region)
            final_dst = os.path.join(RECOGNIZED_SALES_REPORTS_DIR, final_name)

            if os.path.abspath(final_dst) != os.path.abspath(tmp_dst):
                if os.path.exists(final_dst):
                    os.remove(final_dst)
                os.replace(tmp_dst, final_dst)
            else:
                final_dst = tmp_dst

            # Додаємо рядки у зведений results.xlsx (по регіонах)
            if region in {"UA", "KZ", "UZ"}:
                try:
                    headers_src, rows_src = read_cleaned_rows(final_dst)
                    if region not in region_headers:
                        region_headers[region] = headers_src
                        region_rows[region].extend(rows_src)
                    else:
                        aligned = align_rows_to_headers(region_headers[region], headers_src, rows_src)
                        region_rows[region].extend(aligned)
                except Exception as agg_exc:
                    LOGGER.warning("Aggregation skipped for file=%s reason=%s", final_dst, agg_exc)

            llm_flag = "LLM" if used_llm else "H"
            print(
                f"{_fmt_status('OK', use_color)}   {filename:<{FILENAME_COL_WIDTH}} | "
                f"det={llm_flag:<3} | region={region:<3} | checked={checked:<4} | "
                f"mismatches={mism:<4} | saved={os.path.basename(final_dst)}"
            )

        except Exception as exc:
            fail_dst = safe_copy_to_failed(src)
            LOGGER.exception("FAIL file=%s error=%s", src, exc)
            print(f"{_fmt_status('FAIL', use_color)} {filename} | {exc} | copied to {fail_dst}")

    # Формуємо/оновлюємо SALES/RESULT/results.xlsx
    try:
        payload: Dict[str, Tuple[List[object], List[List[object]]]] = {}
        for reg in ("UA", "KZ", "UZ"):
            payload[reg] = (region_headers.get(reg, []), region_rows.get(reg, []))
        results_path = upsert_results_workbook(payload)
        print(f"{_fmt_status('OK', use_color)}   | step=1 | saved={results_path}")
    except Exception as exc:
        LOGGER.exception("Failed to build results.xlsx: %s", exc)
        print(f"{_fmt_status('WRN', use_color)}  | step=1 | results.xlsx_not_built | {exc}")





    summary = (
        f"Проаналізовано: {total_files} файлів | "
        f"Сформовано: {success_files} файлів (із них за допомогою LLM: {llm_used_files})"
    )
    # print()  # порожній рядок для відділення підсумку
    print(summary)
    LOGGER.info(summary)


    return os.path.abspath(RESULTS_XLSX)



def main() -> None:
    """CLI entrypoint: запускає Крок 1 і друкує шлях до results.xlsx."""
    path = build_results_workbook()
    print(f"{_fmt_status('OK', True)}   | step=1 | saved={path}")


if __name__ == "__main__":
    main()